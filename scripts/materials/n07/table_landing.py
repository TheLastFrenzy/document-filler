from dataclasses import dataclass, field
from pathlib import Path
import re
import tempfile
import zipfile

import openpyxl

from materials.shared.embedded_docx import (
    package_stream_from_ole,
    read_ole_anchors,
)
from materials.shared.ledger import merged_value_getter
from materials.shared.ledger_sheet import select_ledger_sheet


REQUIRED_TABLE_LANDING_HEADERS = (
    "服务目录",
    "需求单号",
    "工单号",
    "工单标题",
    "程序数",
    "工单描述",
    "结果表清单",
    "数据统计分析执行周期",
    "数据更新要求",
    "自测报告附件",
    "下发前置机中文名",
)

LEDGER_HEADER_ALIASES = {
    "数据统计分析执行周期": ("执行周期", "数据统计分析执行周期"),
    "自测报告附件": (
        "附件",
        "自测报告附件",
        "03-数据统计分析_测试文档_工单自测报告附件",
    ),
}


def ledger_header_index(headers, canonical_name):
    for alias in LEDGER_HEADER_ALIASES.get(canonical_name, (canonical_name,)):
        if alias in headers:
            return headers.index(alias)
    return -1

REQUIRED_TASK_HEADERS = (
    "落地库名",
    "落地表名",
    "表中文名称",
)

EVIDENCE_TASK_NAME_HEADERS = ("任务名", "调度名", "结果表清单")


@dataclass
class TableLandingTask:
    landing_database: str
    landing_table: str
    business_scene: str
    source_table: str = ""
    launch_time: str = ""
    landing_data_count: str = ""
    dispatch_description: str = ""
    source_volume_image: bytes | None = None
    target_volume_image: bytes | None = None


@dataclass
class TableLandingWorkOrder:
    demand_no: str
    work_order_no: str
    title: str
    description: str
    program_count: int
    source_rows: tuple[int, ...]
    source_tables: list[str]
    target_user: str
    database_type: str
    update_cycle: str
    update_requirement: str
    dispatch_descriptions: list[str] = field(default_factory=list)
    attachment_path: Path | None = None
    tasks: list[TableLandingTask] = field(default_factory=list)


def _parse_program_count(value):
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        raise ValueError(f"程序数不是有效整数: {value}") from None


def _read_order_rows(excel_path, service_dir):
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = select_ledger_sheet(workbook)
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    missing = [header for header in REQUIRED_TABLE_LANDING_HEADERS if ledger_header_index(headers, header) < 0]
    if missing:
        workbook.close()
        raise ValueError(f"台账缺少必要列: {', '.join(missing)}")

    columns = {header: ledger_header_index(headers, header) + 1 for header in REQUIRED_TABLE_LANDING_HEADERS}
    database_type_column = headers.index("下发前置机数据库类型") + 1 if "下发前置机数据库类型" in headers else None
    dispatch_description_column = (
        headers.index("下发任务简单说明（50字以内）") + 1
        if "下发任务简单说明（50字以内）" in headers
        else None
    )
    get = merged_value_getter(sheet)
    grouped = {}
    order_keys = []
    for row in range(2, sheet.max_row + 1):
        if get(row, columns["服务目录"]) != service_dir:
            continue
        work_order_no = get(row, columns["工单号"])
        if not work_order_no:
            workbook.close()
            raise ValueError(f"第{row}行缺少工单号")
        if work_order_no not in grouped:
            grouped[work_order_no] = {
                "demand_no": get(row, columns["需求单号"]),
                "title": get(row, columns["工单标题"]),
                "description": get(row, columns["工单描述"]),
                "program_count": _parse_program_count(get(row, columns["程序数"])),
                "source_rows": [],
                "source_tables": [],
                "target_user": get(row, columns["下发前置机中文名"]),
                "database_type": get(row, database_type_column) if database_type_column else "",
                "update_cycle": get(row, columns["数据统计分析执行周期"]),
                "update_requirement": get(row, columns["数据更新要求"]),
                "dispatch_descriptions": [],
            }
            order_keys.append(work_order_no)
        group = grouped[work_order_no]
        group["source_rows"].append(row)
        source_table = get(row, columns["结果表清单"])
        if source_table:
            group["source_tables"].append(source_table)
            group["dispatch_descriptions"].append(
                get(row, dispatch_description_column) if dispatch_description_column else ""
            )
    workbook.close()

    if not grouped:
        raise ValueError(f"未找到匹配数据: 服务目录={service_dir}")

    return [
        TableLandingWorkOrder(
            demand_no=grouped[key]["demand_no"],
            work_order_no=key,
            title=grouped[key]["title"],
            description=grouped[key]["description"],
            program_count=grouped[key]["program_count"],
            source_rows=tuple(grouped[key]["source_rows"]),
            source_tables=grouped[key]["source_tables"],
            target_user=grouped[key]["target_user"],
            database_type=grouped[key]["database_type"],
            update_cycle=grouped[key]["update_cycle"],
            update_requirement=grouped[key]["update_requirement"],
            dispatch_descriptions=grouped[key]["dispatch_descriptions"],
        )
        for key in order_keys
    ]


def _header_column(excel_path, header):
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet = select_ledger_sheet(workbook)
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    workbook.close()
    index = ledger_header_index(headers, header)
    if index < 0:
        raise ValueError(f"台账缺少必要列: {header}")
    return index + 1


def _workbook_payload(data):
    if data.startswith(b"PK\x03\x04"):
        return data
    return package_stream_from_ole(data)


def extract_embedded_workbooks_by_work_order(excel_path, work_orders, attachment_header, work_dir):
    attachment_column = _header_column(excel_path, attachment_header)
    output_dir = Path(work_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    workbooks = {}
    for anchor in read_ole_anchors(excel_path):
        if anchor.column != attachment_column:
            continue
        order = next((item for item in work_orders if anchor.row in item.source_rows), None)
        if order is None:
            continue
        try:
            payload = _workbook_payload(anchor.payload)
        except Exception as exc:
            raise ValueError(f"工单{order.work_order_no}的自测报告附件无法解析为Excel: {exc}") from exc
        path = output_dir / f"{order.work_order_no}_self_report.xlsx"
        path.write_bytes(payload)
        workbooks[order.work_order_no] = path

    missing = [item.work_order_no for item in work_orders if item.work_order_no not in workbooks]
    if missing:
        raise ValueError(f"以下工单缺少可解析的自测报告附件: {', '.join(missing)}")
    return workbooks


def _extract_images_via_cellimages(excel_path):
    result = {}
    with zipfile.ZipFile(excel_path, "r") as archive:
        if "xl/cellimages.xml" not in archive.namelist():
            return result
        cell_images = archive.read("xl/cellimages.xml").decode("utf-8")
        name_to_rid = {}
        for block in re.findall(r"<etc:cellImage>(.*?)</etc:cellImage>", cell_images, re.DOTALL):
            name_match = re.search(r'name="([^"]+)"', block)
            rid_match = re.search(r'r:embed="(rId\d+)"', block)
            if name_match and rid_match:
                name_to_rid[name_match.group(1)] = rid_match.group(1)
        relationships_path = "xl/_rels/cellimages.xml.rels"
        if relationships_path not in archive.namelist():
            return result
        relationships = archive.read(relationships_path).decode("utf-8")
        rid_to_file = {}
        for match in re.finditer(r'<Relationship[^>]*Id="(rId\d+)"[^>]*Target="([^"]+)"', relationships):
            target = match.group(2)
            if target.startswith("../"):
                target = "xl/" + target[3:]
            elif target.startswith("xl/"):
                target = target
            else:
                target = "xl/" + target.lstrip("/")
            rid_to_file[match.group(1)] = target
        for name, rid in name_to_rid.items():
            target = rid_to_file.get(rid)
            if target:
                result[name] = archive.read(target)
    return result


def _header_index(headers, candidates):
    for candidate in candidates:
        if candidate in headers:
            return headers.index(candidate) + 1
    return None


def _cell_text(sheet, row, column):
    if not column or row > sheet.max_row:
        return ""
    return str(sheet.cell(row, column).value or "").strip()


def _dispimg_image(value, images_by_name):
    match = re.search(r'DISPIMG\("([^"]+)"', str(value or ""))
    if match and match.group(1) in images_by_name:
        return images_by_name[match.group(1)]
    return None


def _evidence_rows_by_row(workbook_path):
    images_by_name = _extract_images_via_cellimages(workbook_path)
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    if len(workbook.worksheets) < 2:
        workbook.close()
        return {}
    sheet = workbook.worksheets[1]
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    columns = {
        "source_table": _header_index(headers, EVIDENCE_TASK_NAME_HEADERS),
        "dispatch_description": _header_index(headers, ("任务中文名", "下发任务简单说明（50字以内）")),
        "landing_database": _header_index(headers, ("落地库名",)),
        "landing_table": _header_index(headers, ("落地表名",)),
        "business_scene": _header_index(headers, ("表中文名", "表中文名称")),
        "source_volume_image": _header_index(headers, ("源表数据量",)),
        "target_volume_image": _header_index(headers, ("目标表数据量",)),
    }
    rows = {}
    for row in range(2, sheet.max_row + 1):
        info = {
            "source_table": _cell_text(sheet, row, columns["source_table"]),
            "dispatch_description": _cell_text(sheet, row, columns["dispatch_description"]),
            "landing_database": _cell_text(sheet, row, columns["landing_database"]),
            "landing_table": _cell_text(sheet, row, columns["landing_table"]),
            "business_scene": _cell_text(sheet, row, columns["business_scene"]),
            "source_volume_image": _dispimg_image(
                _cell_text(sheet, row, columns["source_volume_image"]),
                images_by_name,
            ),
            "target_volume_image": _dispimg_image(
                _cell_text(sheet, row, columns["target_volume_image"]),
                images_by_name,
            ),
        }
        if any(value for value in info.values() if not isinstance(value, bytes)):
            rows[row] = info
    workbook.close()
    return rows


def _target_volume_images_by_row(workbook_path):
    return {
        row: info["target_volume_image"]
        for row, info in _evidence_rows_by_row(workbook_path).items()
        if info.get("target_volume_image")
    }


def _source_key(value):
    return re.sub(r"\s+", "", str(value or "")).strip().upper()


def _order_tasks_by_source_tables(tasks, source_tables):
    if not source_tables or not any(task.source_table for task in tasks):
        return tasks

    by_source = {}
    for task in tasks:
        key = _source_key(task.source_table)
        if key:
            by_source.setdefault(key, []).append(task)

    ordered = []
    missing = []
    for source_table in source_tables:
        matches = by_source.get(_source_key(source_table)) or []
        if matches:
            ordered.append(matches.pop(0))
        else:
            missing.append(source_table)
    if missing:
        raise ValueError(f"附件第二个sheet未匹配到台账结果表清单: {', '.join(missing)}")
    return ordered


def extract_ledger_images_by_row(excel_path, row_numbers, image_columns):
    images_by_name = _extract_images_via_cellimages(excel_path)
    if not images_by_name:
        return {}

    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=False)
    sheet = select_ledger_sheet(workbook)
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    missing = [column for column in image_columns if column not in headers]
    if missing:
        workbook.close()
        raise ValueError(f"台账缺少必要列: {', '.join(missing)}")

    columns = {header: headers.index(header) + 1 for header in image_columns}
    matched = {}
    for row in row_numbers:
        for header, column in columns.items():
            value = str(sheet.cell(row, column).value or "")
            match = re.search(r'DISPIMG\("([^"]+)"', value)
            if match and match.group(1) in images_by_name:
                matched[(row, header)] = images_by_name[match.group(1)]
    workbook.close()
    return matched


def parse_landing_tasks(workbook_path, source_tables=None):
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    missing = [header for header in REQUIRED_TASK_HEADERS if header not in headers]
    if missing:
        workbook.close()
        raise ValueError(f"附件缺少必要列: {', '.join(missing)}")
    columns = {header: headers.index(header) + 1 for header in REQUIRED_TASK_HEADERS}
    landing_data_column = headers.index("落地数据量") + 1 if "落地数据量" in headers else None
    launch_time_column = headers.index("上线时间") + 1 if "上线时间" in headers else None
    dispatch_column = (
        headers.index("下发任务简单说明（50字以内）") + 1
        if "下发任务简单说明（50字以内）" in headers
        else None
    )
    evidence_rows = _evidence_rows_by_row(workbook_path)
    max_row = max([sheet.max_row, *evidence_rows.keys()]) if evidence_rows else sheet.max_row

    tasks = []
    for row in range(2, max_row + 1):
        evidence = evidence_rows.get(row, {})
        landing_database = _cell_text(sheet, row, columns["落地库名"]) or evidence.get("landing_database", "")
        landing_table = _cell_text(sheet, row, columns["落地表名"]) or evidence.get("landing_table", "")
        business_scene = _cell_text(sheet, row, columns["表中文名称"]) or evidence.get("business_scene", "")
        landing_data_count = _cell_text(sheet, row, landing_data_column) if landing_data_column else ""
        launch_time = _cell_text(sheet, row, launch_time_column) if launch_time_column else ""
        dispatch_description = (
            _cell_text(sheet, row, dispatch_column) if dispatch_column else ""
        ) or evidence.get("dispatch_description", "")
        source_table = evidence.get("source_table", "")
        if not any((landing_database, landing_table, business_scene, source_table)):
            continue
        tasks.append(
            TableLandingTask(
                landing_database=landing_database,
                landing_table=landing_table,
                business_scene=business_scene,
                source_table=source_table,
                launch_time=launch_time,
                landing_data_count=landing_data_count,
                dispatch_description=dispatch_description,
                source_volume_image=evidence.get("source_volume_image"),
                target_volume_image=evidence.get("target_volume_image"),
            )
        )
    workbook.close()
    if not tasks:
        raise ValueError(f"附件未解析到库表落地明细: {workbook_path}")
    return _order_tasks_by_source_tables(tasks, source_tables or [])


def read_table_landing_work_orders(excel_path, service_dir):
    orders = _read_order_rows(excel_path, service_dir)
    with tempfile.TemporaryDirectory(prefix="document_filler_table_landing_") as temp_dir:
        workbooks = extract_embedded_workbooks_by_work_order(
            excel_path,
            orders,
            "自测报告附件",
            Path(temp_dir) / "workbooks",
        )
        for order in orders:
            order.attachment_path = workbooks[order.work_order_no]
            order.tasks = parse_landing_tasks(order.attachment_path, order.source_tables)
    return orders
