from dataclasses import dataclass
from pathlib import Path

import openpyxl
from docx import Document
from docx.oxml.ns import qn

from materials.n07.table_landing import read_table_landing_work_orders
from materials.shared.ledger import merged_value_getter
from materials.shared.ledger_sheet import select_ledger_sheet
from materials.shared.word_sections import (
    clone_paragraph_with_text,
    clone_table_with_data,
    element_text,
    elements_between,
    find_heading,
    replace_after,
    replace_between,
    update_toc_via_com,
)


REQUIREMENT_LIST_HEADERS = (
    "服务目录",
    "需求单号",
    "工单号",
    "工单标题",
    "程序数",
)


@dataclass(frozen=True)
class TableLandingRequirementPrototypes:
    business_summary: object
    demand_source_table: object
    work_order_heading: object
    demand_body: object
    task_count: object
    detail_table: object
    supply_method: object
    update_frequency: object
    update_time: object


@dataclass(frozen=True)
class TableLandingRequirementListOrder:
    demand_no: str
    work_order_no: str
    title: str
    program_count: int
    source_rows: tuple[int, ...]


def _paragraph_style_id(element):
    properties = element.find(qn("w:pPr"))
    style = properties.find(qn("w:pStyle")) if properties is not None else None
    return style.get(qn("w:val")) if style is not None else None


def _require_prototype(value, name):
    if value is None:
        raise ValueError(f"模板中未找到{name}格式原型")
    return value


def _find_heading_or_none(document, style_name, exact_text):
    for paragraph in document.paragraphs:
        if paragraph.style.name == style_name and paragraph.text.strip() == exact_text:
            return paragraph
    return None


def _first_nonempty_paragraph(elements, start=0):
    for index in range(start, len(elements)):
        element = elements[index]
        if element.tag == qn("w:p") and element_text(element):
            return element
    return None


def _heading_style_ids(document):
    style_ids = set()
    for level in range(1, 10):
        try:
            style_ids.add(document.styles[f"Heading {level}"].style_id)
        except KeyError:
            continue
    return style_ids


def _elements_until_next_heading(document, heading):
    heading_styles = _heading_style_ids(document)
    elements = []
    current = heading._p.getnext()
    while current is not None and current.tag != qn("w:sectPr"):
        if current.tag == qn("w:p") and _paragraph_style_id(current) in heading_styles and element_text(current):
            return elements, current
        elements.append(current)
        current = current.getnext()
    return elements, None


def _paragraph_with_prefix(elements, prefix):
    for element in elements:
        if element.tag == qn("w:p") and element_text(element).startswith(prefix):
            return element
    return None


def _parse_program_count(value):
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        raise ValueError(f"程序数不是有效整数: {value}") from None


def _read_requirement_list_orders(excel_path, service_dir):
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = select_ledger_sheet(workbook)
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    missing = [header for header in REQUIREMENT_LIST_HEADERS if header not in headers]
    if missing:
        workbook.close()
        raise ValueError(f"台账缺少必要列: {', '.join(missing)}")

    columns = {header: headers.index(header) + 1 for header in REQUIREMENT_LIST_HEADERS}
    get = merged_value_getter(sheet)
    grouped = {}
    order_keys = []
    for row in range(2, sheet.max_row + 1):
        if get(row, columns["服务目录"]) != service_dir:
            continue
        work_order_no = get(row, columns["工单号"])
        if not work_order_no:
            workbook.close()
            raise ValueError(f"第{row}行缺少工单号")
        if work_order_no not in grouped:
            grouped[work_order_no] = {
                "demand_no": get(row, columns["需求单号"]),
                "title": get(row, columns["工单标题"]),
                "program_count": _parse_program_count(get(row, columns["程序数"])),
                "source_rows": [],
            }
            order_keys.append(work_order_no)
        grouped[work_order_no]["source_rows"].append(row)
    workbook.close()

    if not grouped:
        raise ValueError(f"未找到匹配数据: 服务目录={service_dir}")

    return [
        TableLandingRequirementListOrder(
            demand_no=grouped[key]["demand_no"],
            work_order_no=key,
            title=grouped[key]["title"],
            program_count=grouped[key]["program_count"],
            source_rows=tuple(grouped[key]["source_rows"]),
        )
        for key in order_keys
    ]


def _first_table(elements, column_count=None):
    for element in elements:
        if element.tag != qn("w:tbl"):
            continue
        if column_count is None:
            return element
        grid = element.find(qn("w:tblGrid"))
        if grid is not None and len(grid.findall(qn("w:gridCol"))) == column_count:
            return element
        first_row = element.find(qn("w:tr"))
        if first_row is not None and len(first_row.findall(qn("w:tc"))) == column_count:
            return element
    return None


def _build_requirement_list(orders, paragraph_prototype, table_prototype):
    total = sum(order.program_count for order in orders)
    request_count = len({order.demand_no for order in orders if order.demand_no})
    work_order_count = len({order.work_order_no for order in orders if order.work_order_no})
    summary = (
        f"在本服务周期（2025年08月-2026年07月）内，共完成{total}个落地下发，"
        f"涉及{request_count}个需求单、{work_order_count}个工单。"
    )
    headers = ["序号", "对应需求单编号", "对应工单编号", "业务需求内容", "次数"]
    rows = [
        [str(index), order.demand_no, order.work_order_no, order.title, str(order.program_count)]
        for index, order in enumerate(orders, start=1)
    ]
    return [
        clone_paragraph_with_text(paragraph_prototype, summary),
        clone_table_with_data(table_prototype, headers, rows),
    ]


def _fill_redesigned_requirement_list(document, excel_path, service_dir, demand_list_heading):
    orders = _read_requirement_list_orders(excel_path, service_dir)
    section_elements, next_heading = _elements_until_next_heading(document, demand_list_heading)
    paragraph_prototype = _require_prototype(
        _first_nonempty_paragraph(section_elements),
        "需求清单说明段落",
    )
    table_prototype = _require_prototype(
        _first_table(section_elements, 5),
        "需求清单表格",
    )
    replacement = _build_requirement_list(orders, paragraph_prototype, table_prototype)
    if next_heading is None:
        replace_after(demand_list_heading._p, replacement)
    else:
        replace_between(demand_list_heading._p, next_heading, replacement)


def _capture_template_prototypes(document, business_heading, demand_heading):
    business_elements = elements_between(business_heading._p, demand_heading._p)
    demand_source_table = _first_table(business_elements, 5)
    business_summary = _first_nonempty_paragraph(business_elements)

    demand_elements = []
    current = demand_heading._p.getnext()
    while current is not None and current.tag != qn("w:sectPr"):
        demand_elements.append(current)
        current = current.getnext()

    heading_2_style = document.styles["Heading 2"].style_id
    work_order_heading = next(
        (
            item
            for item in demand_elements
            if item.tag == qn("w:p") and _paragraph_style_id(item) == heading_2_style and element_text(item)
        ),
        None,
    )
    start = demand_elements.index(work_order_heading) + 1 if work_order_heading is not None else 0
    after_heading = demand_elements[start:]
    demand_body = _paragraph_with_prefix(after_heading, "需求口径")
    if demand_body is None:
        demand_body = _first_nonempty_paragraph(after_heading)
    task_count = _paragraph_with_prefix(after_heading, "涉及共享任务数量")
    detail_table = _first_table(after_heading, 6)
    supply_method = _paragraph_with_prefix(after_heading, "计划供数方式")
    update_frequency = _paragraph_with_prefix(after_heading, "计划更新频率")
    if update_frequency is None:
        update_frequency = _paragraph_with_prefix(after_heading, "更新频率")
    update_time = _paragraph_with_prefix(after_heading, "更新时间")
    if update_time is None:
        update_time = update_frequency

    return TableLandingRequirementPrototypes(
        business_summary=_require_prototype(business_summary, "业务场景说明段落"),
        demand_source_table=_require_prototype(demand_source_table, "业务场景需求来源表格"),
        work_order_heading=_require_prototype(work_order_heading, "工单标题"),
        demand_body=_require_prototype(demand_body, "需求口径段落"),
        task_count=_require_prototype(task_count, "涉及共享任务数量段落"),
        detail_table=_require_prototype(detail_table, "库表落地明细表格"),
        supply_method=_require_prototype(supply_method, "计划供数方式段落"),
        update_frequency=_require_prototype(update_frequency, "计划更新频率段落"),
        update_time=_require_prototype(update_time, "更新时间段落"),
    )


def _build_business_scene(orders, prototypes):
    request_count = len({order.demand_no for order in orders if order.demand_no})
    total = sum(order.program_count for order in orders)
    summary = (
        f"服务周期内，共有{request_count}张需求单，{len(orders)}张工单涉及{total}个库表落地共享任务。"
        "具体需求单、工单和产出如下表："
    )
    headers = ["序号", "对应需求单编号", "对应工单编号", "工单内容", "涉及共享任务数量（个）"]
    rows = [
        [str(index), order.demand_no, order.work_order_no, order.title, str(order.program_count)]
        for index, order in enumerate(orders, start=1)
    ]
    return [
        clone_paragraph_with_text(prototypes.business_summary, summary),
        clone_table_with_data(
            prototypes.demand_source_table,
            headers,
            rows,
            footer=["总计", "", "", "", str(total)],
        ),
    ]


def _source_table_for_task(order, index):
    if index < len(order.source_tables):
        return order.source_tables[index]
    return order.source_tables[-1] if order.source_tables else ""


def _build_detail_table(order, prototypes):
    headers = ["序号", "落地库", "落地表", "对象用户", "库表来源", "业务场景"]
    rows = [
        [
            str(index + 1),
            task.landing_database,
            task.landing_table,
            order.target_user,
            _source_table_for_task(order, index),
            task.business_scene,
        ]
        for index, task in enumerate(order.tasks)
    ]
    return clone_table_with_data(prototypes.detail_table, headers, rows)


def _build_requirement_sections(orders, prototypes):
    elements = []
    for order in orders:
        elements.append(
            clone_paragraph_with_text(
                prototypes.work_order_heading,
                f"{order.work_order_no}_{order.title}",
            )
        )
        elements.append(
            clone_paragraph_with_text(prototypes.demand_body, f"需求口径：{order.description}")
        )
        elements.append(
            clone_paragraph_with_text(prototypes.task_count, f"涉及共享任务数量：{order.program_count}")
        )
        elements.append(_build_detail_table(order, prototypes))
        elements.append(
            clone_paragraph_with_text(prototypes.supply_method, "计划供数方式：库表下发")
        )
        elements.append(
            clone_paragraph_with_text(prototypes.update_frequency, f"计划更新频率：{order.update_cycle}")
        )
        elements.append(
            clone_paragraph_with_text(prototypes.update_time, f"更新时间：{order.update_requirement}")
        )
    return elements


def build_table_landing_requirement_document(excel_path, service_dir, template_path, output_path):
    document = Document(template_path)
    demand_list_heading = _find_heading_or_none(document, "Heading 2", "需求清单")
    if demand_list_heading is not None:
        _fill_redesigned_requirement_list(document, excel_path, service_dir, demand_list_heading)
    else:
        orders = read_table_landing_work_orders(excel_path, service_dir)
        business_heading = find_heading(document, "Heading 2", "业务场景")
        demand_heading = find_heading(document, "Heading 2", "需求说明")
        prototypes = _capture_template_prototypes(document, business_heading, demand_heading)

        replace_between(
            business_heading._p,
            demand_heading._p,
            _build_business_scene(orders, prototypes),
        )
        replace_after(demand_heading._p, _build_requirement_sections(orders, prototypes))

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    document.save(output)
    update_toc_via_com(output)
    return str(output)
