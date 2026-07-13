from dataclasses import dataclass
import os
import posixpath
import tempfile
from pathlib import Path
from xml.etree import ElementTree as ET
import zipfile

import olefile
import openpyxl


REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
OLE_NS = "urn:schemas-microsoft-com:office:office"


@dataclass(frozen=True)
class OleAnchor:
    row: int
    column: int
    payload: bytes


def _local_name(element):
    return element.tag.rsplit("}", 1)[-1]


def _relationship_targets(archive, sheet_path):
    sheet_name = posixpath.basename(sheet_path)
    rels_path = posixpath.join(posixpath.dirname(sheet_path), "_rels", f"{sheet_name}.rels")
    if rels_path not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read(rels_path))
    base = posixpath.dirname(sheet_path)
    return {
        element.attrib["Id"]: posixpath.normpath(posixpath.join(base, element.attrib["Target"]))
        for element in root
        if "Id" in element.attrib and "Target" in element.attrib
    }


def read_ole_anchors(excel_path):
    anchors = []
    with zipfile.ZipFile(excel_path, "r") as archive:
        names = set(archive.namelist())
        sheet_paths = sorted(
            name for name in names if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        for sheet_path in sheet_paths:
            root = ET.fromstring(archive.read(sheet_path))
            targets = _relationship_targets(archive, sheet_path)
            shape_targets = {}
            legacy_rids = []
            for element in root.iter():
                local = _local_name(element)
                if local == "oleObject":
                    shape_id = element.attrib.get("shapeId", "")
                    rid = element.attrib.get(f"{{{REL_NS}}}id", "")
                    if shape_id and rid:
                        shape_targets[shape_id] = targets.get(rid, "")
                elif local == "legacyDrawing":
                    rid = element.attrib.get(f"{{{REL_NS}}}id", "")
                    if rid:
                        legacy_rids.append(rid)

            for legacy_rid in legacy_rids:
                vml_path = targets.get(legacy_rid, "")
                if not vml_path or vml_path not in names:
                    continue
                vml_root = ET.fromstring(archive.read(vml_path))
                for shape in vml_root.iter():
                    if _local_name(shape) != "shape" or shape.attrib.get(f"{{{OLE_NS}}}ole") != "t":
                        continue
                    shape_id = shape.attrib.get("id", "").rsplit("_s", 1)[-1]
                    target = shape_targets.get(shape_id, "")
                    if not target or target not in names:
                        continue
                    anchor_text = ""
                    for child in shape.iter():
                        if _local_name(child) == "Anchor":
                            anchor_text = child.text or ""
                            break
                    parts = [int(value.strip()) for value in anchor_text.split(",") if value.strip()]
                    if len(parts) < 3:
                        continue
                    anchors.append(
                        OleAnchor(
                            row=parts[2] + 1,
                            column=parts[0] + 1,
                            payload=archive.read(target),
                        )
                    )
    return anchors


def package_stream_from_ole(data):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".bin") as temp_file:
        temp_file.write(data)
        temp_path = Path(temp_file.name)
    ole = None
    try:
        ole = olefile.OleFileIO(str(temp_path))
        streams = {
            "/".join(parts).lower(): "/".join(parts)
            for parts in ole.listdir(streams=True, storages=False)
        }
        if "package" in streams:
            payload = ole.openstream(streams["package"]).read()
        else:
            native_name = next((name for key, name in streams.items() if "ole10native" in key), "")
            if not native_name:
                raise ValueError("OLE对象缺少Package或Ole10Native流")
            native = ole.openstream(native_name).read()
            offset = native.find(b"PK\x03\x04")
            if offset < 0:
                raise ValueError("Ole10Native流中未找到DOCX载荷")
            payload = native[offset:]
        if not payload.startswith(b"PK\x03\x04"):
            raise ValueError("内嵌附件不是DOCX文件")
        return payload
    finally:
        if ole is not None:
            ole.close()
        try:
            os.unlink(temp_path)
        except FileNotFoundError:
            pass


def _header_column(excel_path, header):
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    workbook.close()
    if header not in headers:
        raise ValueError(f"台账缺少必要列: {header}")
    return headers.index(header) + 1


def extract_embedded_docx_by_work_order(excel_path, work_orders, attachment_header, work_dir):
    attachment_column = _header_column(excel_path, attachment_header)
    output_dir = Path(work_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    reports = {}
    for anchor in read_ole_anchors(excel_path):
        if anchor.column != attachment_column:
            continue
        order = next((item for item in work_orders if anchor.row in item.source_rows), None)
        if order is None:
            continue
        try:
            payload = package_stream_from_ole(anchor.payload)
        except Exception as exc:
            raise ValueError(f"工单{order.work_order_no}的自测报告附件无法解析: {exc}") from exc
        path = output_dir / f"{order.work_order_no}_self_report.docx"
        path.write_bytes(payload)
        reports[order.work_order_no] = path
    missing = [item.work_order_no for item in work_orders if item.work_order_no not in reports]
    if missing:
        raise ValueError(f"以下工单缺少可解析的自测报告附件: {', '.join(missing)}")
    return reports
