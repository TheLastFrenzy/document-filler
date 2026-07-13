from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


REQUIRED_API_HEADERS = (
    "服务目录",
    "需求单号",
    "工单号",
    "工单标题",
    "程序数",
    "工单描述",
    "结果表清单",
    "自测报告附件",
)


@dataclass
class ParameterGroup:
    label: str
    headers: list[str]
    rows: list[list[str]]


@dataclass
class ApiInterface:
    chinese_name: str
    english_name: str
    source_row: int
    input_groups: list[ParameterGroup] = field(default_factory=list)
    output_groups: list[ParameterGroup] = field(default_factory=list)
    purpose: str = ""


@dataclass
class ApiWorkOrder:
    demand_no: str
    work_order_no: str
    title: str
    description: str
    program_count: int
    source_rows: tuple[int, ...]
    interfaces: list[ApiInterface]
    self_report_path: Path | None = None


def merged_value_getter(sheet):
    values = {}
    for area in sheet.merged_cells.ranges:
        value = sheet.cell(area.min_row, area.min_col).value
        for row in range(area.min_row, area.max_row + 1):
            for column in range(area.min_col, area.max_col + 1):
                values[(row, column)] = value

    def get(row, column):
        value = sheet.cell(row, column).value
        if value is None:
            value = values.get((row, column))
        return "" if value is None else str(value).strip()

    return get


def parse_interface_name(value):
    parts = str(value or "").strip().rsplit(None, 1)
    if len(parts) != 2 or not all(parts):
        raise ValueError(f"结果表清单格式错误，应为‘接口中文名 接口英文名’: {value}")
    return parts[0].strip(), parts[1].strip()


def _parse_program_count(value):
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        raise ValueError(f"程序数不是有效整数: {value}") from None


def read_api_work_orders(excel_path, service_dir):
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = workbook.active
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    missing = [header for header in REQUIRED_API_HEADERS if header not in headers]
    if missing:
        workbook.close()
        raise ValueError(f"台账缺少必要列: {', '.join(missing)}")

    columns = {header: headers.index(header) + 1 for header in REQUIRED_API_HEADERS}
    get = merged_value_getter(sheet)
    grouped = {}
    order_keys = []
    for row in range(2, sheet.max_row + 1):
        if get(row, columns["服务目录"]) != service_dir:
            continue
        work_order_no = get(row, columns["工单号"])
        if not work_order_no:
            workbook.close()
            raise ValueError(f"第{row}行缺少工单号")
        if work_order_no not in grouped:
            grouped[work_order_no] = {
                "demand_no": get(row, columns["需求单号"]),
                "title": get(row, columns["工单标题"]),
                "description": get(row, columns["工单描述"]),
                "program_count": _parse_program_count(get(row, columns["程序数"])),
                "source_rows": [],
                "interfaces": [],
            }
            order_keys.append(work_order_no)
        group = grouped[work_order_no]
        group["source_rows"].append(row)
        result_text = get(row, columns["结果表清单"])
        if result_text:
            chinese_name, english_name = parse_interface_name(result_text)
            if not any(item.chinese_name == chinese_name for item in group["interfaces"]):
                group["interfaces"].append(ApiInterface(chinese_name, english_name, row))
    workbook.close()

    if not grouped:
        raise ValueError(f"未找到匹配数据: 服务目录={service_dir}")

    return [
        ApiWorkOrder(
            demand_no=grouped[key]["demand_no"],
            work_order_no=key,
            title=grouped[key]["title"],
            description=grouped[key]["description"],
            program_count=grouped[key]["program_count"],
            source_rows=tuple(grouped[key]["source_rows"]),
            interfaces=grouped[key]["interfaces"],
        )
        for key in order_keys
    ]
