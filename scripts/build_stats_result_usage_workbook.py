#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the N08 data-statistics result-table workbook from the adjusted ledger."""

from __future__ import annotations

import html
import json
import math
import re
import sys
import textwrap
import zipfile
import xml.etree.ElementTree as ET
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


DEFAULT_SERVICE_DIR = "N08-数据统计分析"
LEDGER_XML_COL = "程序XML文本"
LEDGER_RESULT_COL = "统计分析结果表清单"
LEDGER_COLUMN_ALIASES = {
    LEDGER_RESULT_COL: (LEDGER_RESULT_COL, "结果表清单"),
    "工单内容": ("工单内容", "工单标题"),
    "报表统计次数": ("报表统计次数", "程序数"),
    "业务说明": ("业务说明", "工单描述"),
    "业务描述": ("业务描述", "工单描述", "业务说明"),
}

SHEET_SOURCE_LIST = "1、数据源表list"
SHEET_RELATION = "2、表融合关系"
SHEET_RESULT_LIST = "3、数据统计分析结果表list"
SHEET_RESULT_DETAIL = "4、数据统计分析结果表详情"

NS_MAIN = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_REL = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}

DEFAULT_FONT_NAME = "宋体"
DEFAULT_HIGHLIGHT_FILL = "F2F2F2"
DEFAULT_DATA_PROVIDER = "上海市大数据中心"
DEFAULT_RESOURCE_TYPE = "库表"

STATS_LOGIC_TERM_REPLACEMENTS = (
    ("数据筛选清洗", "数据范围确认"),
    ("筛选清洗", "范围确认"),
    ("清洗加工后的数据", "整理后的数据"),
    ("清洗加工后", "整理后"),
    ("清洗加工", "整理"),
    ("标准化抽取", "按字段口径整理"),
    ("数据抽取", "数据取用"),
    ("抽取", "取用"),
    ("质量检查", "口径核对"),
    ("清洗", "整理"),
    ("加密", ""),
)


def sanitize_stats_logic_text(value: object) -> str:
    text = str(value or "")
    for old, new in STATS_LOGIC_TERM_REPLACEMENTS:
        text = text.replace(old, new)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"（\s*）", "", text)
    text = re.sub(r"：\s*，", "：", text)
    text = re.sub(r"，{2,}", "，", text)
    return text.strip()


def norm_name(value: object) -> str:
    text = str(value or "").strip().strip("`\"[];").upper()
    if "." in text:
        text = text.split(".")[-1]
    text = re.sub(r"\$\{[^}]+\}", "", text)
    return text.strip("_ `\"[];")


def normalize_ledger_record(record: dict[str, str]) -> dict[str, str]:
    normalized = dict(record)
    for canonical_name, aliases in LEDGER_COLUMN_ALIASES.items():
        if str(normalized.get(canonical_name, "") or "").strip():
            continue
        for alias in aliases:
            value = str(normalized.get(alias, "") or "").strip()
            if value:
                normalized[canonical_name] = value
                break
        normalized.setdefault(canonical_name, "")
    return normalized


def base_resource_name(value: object) -> str:
    text = norm_name(value)
    return text


def merged_value_getter(ws):
    merged: dict[tuple[int, int], object] = {}
    for area in ws.merged_cells.ranges:
        value = ws.cell(area.min_row, area.min_col).value
        for row in range(area.min_row, area.max_row + 1):
            for col in range(area.min_col, area.max_col + 1):
                merged[(row, col)] = value

    def get(row: int, col: int) -> str:
        value = merged.get((row, col), ws.cell(row, col).value)
        return "" if value is None else str(value).strip()

    return get


def load_ledger_rows(ledger_path: Path, service_dir: str) -> list[dict[str, str]]:
    wb = load_workbook(ledger_path, data_only=False)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    get = merged_value_getter(ws)
    service_col = headers.index("服务目录") + 1
    rows: list[dict[str, str]] = []
    for row in range(2, ws.max_row + 1):
        if get(row, service_col) != service_dir:
            continue
        record = {header: get(row, idx + 1) for idx, header in enumerate(headers) if header}
        record = normalize_ledger_record(record)
        record["_row"] = str(row)
        rows.append(record)
    return rows


def parse_result_name(text: str) -> tuple[str, str]:
    parts = (text or "").strip().rsplit(None, 1)
    if len(parts) == 2:
        return parts[0].strip(), norm_name(parts[1])
    return text.strip(), ""


def parse_model_data(raw: str | None) -> dict:
    if not raw:
        return {}
    try:
        value = json.loads(html.unescape(raw))
    except Exception:
        return {}
    return value if isinstance(value, dict) else {}


def parse_xml_program(xml_text: str) -> tuple[list[dict], list[dict]]:
    root = ET.fromstring(xml_text)
    nodes: list[dict] = []
    edges: list[dict] = []
    for elem in root.iter():
        if not elem.tag.endswith("mxCell"):
            continue
        if elem.attrib.get("edge") == "1":
            edges.append(
                {
                    "source": elem.attrib.get("source"),
                    "target": elem.attrib.get("target"),
                    "label": elem.attrib.get("value", ""),
                }
            )
            continue
        data = parse_model_data(elem.attrib.get("modelData"))
        label = data.get("stepLabel") or elem.attrib.get("title") or elem.attrib.get("value") or ""
        sql = str(data.get("sql") or "")
        expression = str(data.get("expression") or "")
        if label or sql or expression:
            nodes.append(
                {
                    "id": elem.attrib.get("id", ""),
                    "label": label,
                    "sql": sql,
                    "expression": expression,
                    "style": elem.attrib.get("style", ""),
                }
            )
    return nodes, edges


def strip_sql_comments(sql: str) -> str:
    return re.sub(r"--.*", " ", re.sub(r"/\*.*?\*/", " ", sql, flags=re.S))


def cte_names(sql: str) -> set[str]:
    clean = strip_sql_comments(sql)
    return {
        norm_name(match.group(1))
        for match in re.finditer(r"(?:\bwith|,)\s+([A-Za-z_][\w]*)\s+as\s*\(", clean, flags=re.I)
    }


def source_tables(sqls: list[str], target: str) -> list[str]:
    ctes: set[str] = set()
    for sql in sqls:
        ctes |= cte_names(sql)

    sources: list[str] = []
    for sql in sqls:
        clean = strip_sql_comments(sql)
        for pattern in (r"\bfrom\s+([`\"\w\.\$\{\}]+)", r"\bjoin\s+([`\"\w\.\$\{\}]+)"):
            for match in re.finditer(pattern, clean, flags=re.I):
                name = base_resource_name(match.group(1))
                if (
                    not name
                    or name in {target, "ALL_TABLES", "DUAL", "SELECT"}
                    or name in ctes
                    or name.startswith("TEMP_")
                    or len(name) <= 2
                ):
                    continue
                if name not in sources:
                    sources.append(name)
    return sources


def find_create_body(sql: str, target: str, allow_fallback: bool = False) -> tuple[str, str] | tuple[None, None]:
    pattern = re.compile(
        r"create\s+(?:external\s+)?table\s+(?:if\s+not\s+exists\s+)?[`\"]?([A-Za-z_][\w]*(?:\.[A-Za-z_][\w]*)?)[`\"]?",
        flags=re.I,
    )
    for match in pattern.finditer(sql):
        table_name = norm_name(match.group(1))
        target_match = table_name == target or table_name == f"{target}Y"
        if not target_match and not allow_fallback:
            continue
        pos = sql.find("(", match.end())
        if pos < 0:
            continue
        depth = 0
        quote: str | None = None
        for idx in range(pos, len(sql)):
            ch = sql[idx]
            if quote:
                if ch == quote:
                    quote = None
                continue
            if ch in "'\"`":
                quote = ch
                continue
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    return sql[pos + 1 : idx], table_name
    return None, None


def split_columns(body: str) -> list[str]:
    out: list[str] = []
    current = ""
    depth = 0
    quote: str | None = None
    for ch in body:
        if quote:
            current += ch
            if ch == quote:
                quote = None
            continue
        if ch in "'\"`":
            quote = ch
            current += ch
            continue
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth = max(0, depth - 1)
        elif ch == "," and depth == 0:
            if current.strip():
                out.append(current.strip())
            current = ""
            continue
        current += ch
    if current.strip():
        out.append(current.strip())
    return out


def oracle_column_comments(sqls: list[str], target: str) -> dict[str, str]:
    comments: dict[str, str] = {}
    combined = "\n".join(sqls)
    pattern = r"comment\s+on\s+column\s+([`\"\w\.]+)\.([`\"\w]+)\s+is\s+'([^']*)'"
    for match in re.finditer(pattern, combined, flags=re.I):
        if norm_name(match.group(1)) == target:
            comments[match.group(2).strip("`\"").upper()] = match.group(3)
    return comments


def strip_leading_sql_comments(text: str) -> str:
    value = text.strip()
    while True:
        next_value = re.sub(r"^\s*--[^\r\n]*(?:\r?\n|$)", "", value, count=1)
        next_value = re.sub(r"^\s*/\*.*?\*/", "", next_value, count=1, flags=re.S)
        next_value = next_value.strip()
        if next_value == value:
            return value
        value = next_value


def parse_result_fields_from_sql(sqls: list[str], target: str) -> list[dict[str, str]]:
    body = None
    body_table = None
    for sql in sqls:
        body, body_table = find_create_body(sql, target)
        if body:
            break
    if not body and len(sqls) <= 3:
        for sql in sqls:
            body, body_table = find_create_body(sql, target, allow_fallback=True)
            if body:
                break
    if not body:
        return []

    oracle_comments = oracle_column_comments(sqls, body_table or target)
    fields: list[dict[str, str]] = []
    for raw in split_columns(body):
        raw = strip_leading_sql_comments(raw)
        if re.match(r"^(primary|unique|constraint|key|index|partitioned|stored)\b", raw, flags=re.I):
            continue
        match = re.match(r"[`\"]?([A-Za-z_][\w]*)[`\"]?\s+(.*)$", raw, flags=re.S)
        if not match:
            continue
        field_en = match.group(1).upper()
        rest = match.group(2).strip()
        inline_comment = ""
        comment_match = re.search(r"\bcomment\s+'([^']*)'", rest, flags=re.I)
        if comment_match:
            inline_comment = comment_match.group(1)
        default = "NULL"
        default_match = re.search(r"\bdefault\s+([^\s,]+)", rest, flags=re.I)
        if default_match:
            default = default_match.group(1)
        nullable = "Yes" if re.search(r"\bnot\s+null\b", rest, flags=re.I) else "No"
        unique = "Yes" if re.search(r"\bunique\b|\bprimary\s+key\b", rest, flags=re.I) else "No"
        field_type = re.split(
            r"\bdefault\b|\bnot\s+null\b|\bunique\b|\bprimary\s+key\b|\bcomment\b",
            rest,
            flags=re.I,
        )[0].strip()
        fields.append(
            {
                "字段中文名": oracle_comments.get(field_en) or inline_comment or field_en,
                "字段英文名": field_en,
                "字段类型": field_type,
                "默认": default,
                "不可为空": nullable,
                "唯一": unique,
                "字段注释": "",
                "来源": "XML",
            }
        )
    return fields


def sql_comment_summary(sql: str) -> str:
    comments = []
    for line in sql.splitlines():
        stripped = line.strip()
        if stripped.startswith("--"):
            comment = stripped[2:].strip()
            if comment:
                comments.append(comment)
    if comments:
        return "；".join(comments[:2])
    if re.search(r"\bcreate\s+table\b", sql, flags=re.I):
        return "创建结果表或中间加工表"
    if re.search(r"\binsert\s+into\b", sql, flags=re.I):
        return "将整理后的数据写入目标表"
    if re.search(r"\btruncate\s+table\b", sql, flags=re.I):
        return "清空目标表历史数据"
    if re.search(r"\balter\s+table\b", sql, flags=re.I):
        return "维护分区或表结构"
    first = " ".join(sql.strip().split())
    return first[:80] or "执行程序节点"


def compact_join(items: list[str], limit: int = 4) -> str:
    if not items:
        return ""
    shown = items[:limit]
    suffix = f" 等{len(items)}张表" if len(items) > limit else ""
    return "、".join(shown) + suffix


def source_label(source: str, resource_info: dict[str, dict] | None = None) -> str:
    info = (resource_info or {}).get(source, {})
    resource_name = str(info.get("资源名称", "") or "").strip()
    if resource_name and resource_name != source:
        return sanitize_stats_logic_text(f"{resource_name}（{source}）")
    return source


def result_theme(result_cn: str, target: str) -> str:
    text = f"{result_cn} {target}".lower()
    if "环比" in text:
        return "工单环比变化"
    if "收单" in text or "总量" in text or "汇总" in text:
        return "工单数量汇总"
    if "分类" in text or "饼图" in text:
        return "工单分类统计"
    if "问题" in text or "下钻" in text:
        return "问题明细分析"
    if "应急" in text:
        return "应急处置专题分析"
    if "妇联" in text or "救助" in text:
        return "困难人员救助名单分析"
    if "高新" in text or "社保" in text:
        return "企业与人员资质核验分析"
    if "12345" in text or "热线" in text:
        return "热线工单与回访数据分析"
    return "统计分析结果"


def extract_where_hints(sqls: list[str]) -> list[str]:
    combined = "\n".join(sqls)
    hints: list[str] = []
    if re.search(r"max\s*\(\s*bdc_d[tt]\s*\)|max\s*\(\s*bdc_mt\s*\)", combined, flags=re.I):
        hints.append("取最新分区/最新批次数据")
    if re.search(r"jhpt_delete\s*=\s*0|delete\s*=\s*0|未删除", combined, flags=re.I):
        hints.append("剔除已删除或无效记录")
    if re.search(r"service_id|服务ID", combined, flags=re.I):
        hints.append("限定指定服务或专题范围")
    if re.search(r"row_number\s*\(\s*\)\s*over", combined, flags=re.I):
        hints.append("按业务主键和更新时间去重取最新记录")
    if re.search(r"corp_status|status\s*=\s*'有效'|有效企业", combined, flags=re.I):
        hints.append("保留有效企业/有效状态记录")
    if re.search(r"addr_area_id_gb|浦东", combined, flags=re.I):
        hints.append("限定行政区划或区域条件")
    if re.search(r"aae140|aae150|aae180|社保", combined, flags=re.I):
        hints.append("按社保缴费险种、年月或缴费基数筛选")
    if re.search(r"12345|热线|工单", combined, flags=re.I):
        hints.append("围绕热线工单、回访或办结记录提取数据")
    if re.search(r"bdc_dt\s*=\s*'\$\{taskid\}'|partition\s*\(", combined, flags=re.I):
        hints.append("按任务批次或分区落表")
    return list(dict.fromkeys(hints))


def extract_calc_hints(sqls: list[str]) -> list[str]:
    combined = "\n".join(sqls)
    hints: list[str] = []
    if re.search(r"\bcount\s*\(", combined, flags=re.I):
        hints.append("统计数量类指标")
    if re.search(r"\bgroup\s+by\b", combined, flags=re.I):
        hints.append("按业务维度分组汇总")
    if re.search(r"\bleft\s+join\b|\binner\s+join\b|\bjoin\b", combined, flags=re.I):
        hints.append("按主键或业务编码关联补齐维度信息")
    if re.search(r"\bunion\b", combined, flags=re.I):
        hints.append("合并多来源记录")
    if re.search(r"\bcase\b|\bif\s*\(", combined, flags=re.I):
        hints.append("按条件生成分类、标签或状态字段")
    if re.search(r"环比|同比|/\s*[^\\n]*\*\s*100|percent|rate", combined, flags=re.I):
        hints.append("计算环比、同比或占比类指标")
    if re.search(r"distinct", combined, flags=re.I):
        hints.append("去重形成唯一结果记录")
    return list(dict.fromkeys(hints))


def build_business_logic_steps(record: dict, resource_info: dict[str, dict]) -> list[str]:
    result_cn = record["result_cn"]
    result_en = record["result_en"]
    sources = record["sources"]
    sqls = [node["sql"] for node in record["nodes"] if node.get("sql")]
    source_text = compact_join([source_label(src, resource_info) for src in sources], limit=3)
    theme = result_theme(result_cn, result_en)
    where_hints = extract_where_hints(sqls)
    calc_hints = extract_calc_hints(sqls)

    steps: list[str] = []
    if source_text:
        steps.append(f"数据来源准备：读取{source_text}，作为{theme}的基础数据。")
    else:
        steps.append(f"数据来源准备：读取程序中配置的基础数据，作为{theme}的输入。")

    if where_hints:
        steps.append(f"数据范围确认：{ '，'.join(where_hints[:4]) }，形成可参与统计分析的业务记录范围。")
    else:
        steps.append("数据范围确认：按结果表字段口径整理源表记录，保留本次统计分析所需的业务范围。")

    if calc_hints:
        steps.append(f"融合加工计算：{ '，'.join(calc_hints[:4]) }，生成结果表所需的统计口径和明细字段。")
    else:
        steps.append("融合加工计算：按照结果表字段映射关系整理数据，形成可直接落表的统计分析结果。")

    field_names = [field.get("字段中文名", "") for field in record.get("fields", []) if field.get("字段中文名")]
    if field_names:
        preview = "、".join(field_names[:4])
        suffix = "等字段" if len(field_names) > 4 else "字段"
        steps.append(f"结果字段整理：输出{preview}{suffix}，补充业务时间、批次号等运行信息。")

    steps.append(f"结果输出：将统计结果写入{result_cn}（{result_en}），供后续数据统计分析应用使用。")
    return [sanitize_stats_logic_text(step) for step in steps]


def xml_sheet_paths(xlsx_path: Path) -> dict[str, str]:
    with zipfile.ZipFile(xlsx_path) as zf:
        workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rels_xml
        }
        mapping: dict[str, str] = {}
        for sheet in workbook_xml.find("a:sheets", NS_MAIN):
            name = sheet.attrib["name"]
            rid = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
            target = rid_to_target[rid]
            if not target.startswith("xl/"):
                target = "xl/" + target
            mapping[name] = target
        return mapping


def load_shared_strings(xlsx_path: Path) -> list[str]:
    strings: list[str] = []
    with zipfile.ZipFile(xlsx_path) as zf:
        if "xl/sharedStrings.xml" not in zf.namelist():
            return strings
        with zf.open("xl/sharedStrings.xml") as fh:
            for _, elem in ET.iterparse(fh, events=("end",)):
                if elem.tag.endswith("}si"):
                    texts = [node.text or "" for node in elem.iter() if node.tag.endswith("}t")]
                    strings.append("".join(texts))
                    elem.clear()
    return strings


def cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        v = cell.find("a:v", NS_MAIN)
        if v is None or v.text is None:
            return ""
        idx = int(v.text)
        return shared_strings[idx] if idx < len(shared_strings) else ""
    if cell_type == "inlineStr":
        texts = [node.text or "" for node in cell.iter() if node.tag.endswith("}t")]
        return "".join(texts)
    v = cell.find("a:v", NS_MAIN)
    return "" if v is None or v.text is None else str(v.text)


def col_index_from_ref(ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", ref.upper())
    total = 0
    for ch in letters:
        total = total * 26 + (ord(ch) - ord("A") + 1)
    return total


def iter_sheet_rows(xlsx_path: Path, sheet_name: str):
    paths = xml_sheet_paths(xlsx_path)
    shared_strings = load_shared_strings(xlsx_path)
    sheet_path = paths[sheet_name]
    with zipfile.ZipFile(xlsx_path) as zf:
        with zf.open(sheet_path) as fh:
            for _, row in ET.iterparse(fh, events=("end",)):
                if not row.tag.endswith("}row"):
                    continue
                values: dict[int, str] = {}
                for cell in row:
                    if not cell.tag.endswith("}c"):
                        continue
                    ref = cell.attrib.get("r", "")
                    if not ref:
                        continue
                    values[col_index_from_ref(ref)] = cell_text(cell, shared_strings)
                max_col = max(values) if values else 0
                yield [values.get(idx, "") for idx in range(1, max_col + 1)]
                row.clear()


def load_resource_info(catalog_path: Path, needed_resources: set[str]) -> tuple[dict[str, dict], set[str]]:
    rows = iter_sheet_rows(catalog_path, "关联资源信息")
    headers = next(rows)
    idx = {name: pos for pos, name in enumerate(headers)}
    info: dict[str, dict] = {}
    needed_dir_codes: set[str] = set()
    for row in rows:
        if len(row) <= idx["资源编码"]:
            continue
        resource_code = norm_name(row[idx["资源编码"]])
        if resource_code in needed_resources and resource_code not in info:
            record = {name: (row[pos] if pos < len(row) else "") for name, pos in idx.items()}
            info[resource_code] = record
            if record.get("数据目录代码"):
                needed_dir_codes.add(str(record["数据目录代码"]).strip())
    return info, needed_dir_codes


def load_catalog_fields(catalog_path: Path, needed_dir_codes: set[str]) -> dict[str, list[dict[str, str]]]:
    if not needed_dir_codes:
        return {}
    rows = iter_sheet_rows(catalog_path, "数据项")
    headers = next(rows)
    idx = {name: pos for pos, name in enumerate(headers)}
    fields: dict[str, list[dict[str, str]]] = {code: [] for code in needed_dir_codes}
    for row in rows:
        code = str(row[idx["数据目录代码"]]).strip() if idx["数据目录代码"] < len(row) else ""
        if code not in needed_dir_codes:
            continue
        fields[code].append(
            {
                "字段中文名": row[idx["数据项名称"]] if idx["数据项名称"] < len(row) else "",
                "字段英文名": row[idx["英文名称"]] if idx["英文名称"] < len(row) else "",
                "字段类型": row[idx["数据类型"]] if idx["数据类型"] < len(row) else "",
                "默认": "NULL",
                "不可为空": "No",
                "唯一": "No",
                "字段注释": row[idx["字段描述"]] if idx["字段描述"] < len(row) else "",
                "来源": "数据目录",
            }
        )
    return fields


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def style_range(ws, min_row: int, max_row: int, min_col: int, max_col: int, template_row: int | None = None) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            if template_row:
                copy_cell_style(ws.cell(template_row, col), cell)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def clear_sheet(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    for image in list(getattr(ws, "_images", [])):
        ws._images.remove(image)
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def setup_common_ws(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def set_widths(ws, widths: list[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_header_row(ws, headers: list[str], row: int = 1) -> None:
    fill = PatternFill("solid", fgColor=DEFAULT_HIGHLIGHT_FILL)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = Font(name=DEFAULT_FONT_NAME, size=11, bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    style_range(ws, row, row, 1, len(headers))


def font_path() -> str | None:
    candidates = [
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\simhei.ttf",
        r"C:\Windows\Fonts\simsun.ttc",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return candidate
    return None


def wrap_by_width(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    def text_w(value: str) -> int:
        bbox = draw.textbbox((0, 0), value, font=font)
        return bbox[2] - bbox[0]

    def char_wrap(raw_text: str) -> list[str]:
        wrapped: list[str] = []
        current = ""
        for ch in raw_text:
            candidate = current + ch
            if text_w(candidate) <= max_width:
                current = candidate
            else:
                if current:
                    wrapped.append(current)
                current = ch
        if current:
            wrapped.append(current)
        return wrapped

    def identifier_wrap(raw_text: str) -> list[str]:
        parts = raw_text.split("_")
        if len(parts) <= 1:
            return char_wrap(raw_text)
        wrapped: list[str] = []
        current = parts[0]
        for part in parts[1:]:
            candidate = f"{current}_{part}" if current else part
            if text_w(candidate) <= max_width:
                current = candidate
            else:
                if current:
                    wrapped.extend(char_wrap(current) if text_w(current) > max_width else [current])
                current = part
        if current:
            wrapped.extend(char_wrap(current) if text_w(current) > max_width else [current])
        return wrapped

    lines: list[str] = []
    for raw in str(text).splitlines() or [""]:
        raw = raw.strip()
        if not raw:
            lines.append("")
            continue
        if draw.textbbox((0, 0), raw, font=font)[2] <= max_width:
            lines.append(raw)
        elif "_" in raw:
            lines.extend(identifier_wrap(raw))
        else:
            lines.extend(char_wrap(raw))
    return lines


def draw_node(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    text: str,
    font: ImageFont.ImageFont,
    fill: str = "#F6D1AD",
    outline: str = "#4B5563",
    rounded: bool = False,
) -> None:
    if rounded:
        draw.rounded_rectangle(box, radius=22, fill=fill, outline=outline, width=3)
    else:
        draw.rectangle(box, fill=fill, outline=outline, width=3)
    max_width = box[2] - box[0] - 24
    lines = wrap_by_width(draw, text, font, max_width)
    visible_count = min(len(lines), 6)
    line_height = min(25, max(18, (box[3] - box[1] - 12) // max(visible_count, 1)))
    total_h = visible_count * line_height
    y = box[1] + (box[3] - box[1] - total_h) / 2
    for line in lines[:visible_count]:
        bbox = draw.textbbox((0, 0), line, font=font)
        x = box[0] + (box[2] - box[0] - (bbox[2] - bbox[0])) / 2
        draw.text((x, y), line, font=font, fill="#333333")
        y += line_height


def draw_arrow(draw: ImageDraw.ImageDraw, start: tuple[int, int], end: tuple[int, int], color: str = "#2F2F2F") -> None:
    draw.line([start, end], fill=color, width=4)
    angle = math.atan2(end[1] - start[1], end[0] - start[0])
    size = 14
    p1 = (end[0] - size * math.cos(angle - math.pi / 6), end[1] - size * math.sin(angle - math.pi / 6))
    p2 = (end[0] - size * math.cos(angle + math.pi / 6), end[1] - size * math.sin(angle + math.pi / 6))
    draw.polygon([end, p1, p2], fill=color)


def draw_flowchart_png(record: dict, output_path: Path) -> None:
    width, height = 1280, 1120
    img = Image.new("RGBA", (width, height), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)
    fp = font_path()
    font_node = ImageFont.truetype(fp, 24) if fp else ImageFont.load_default()
    font_small = ImageFont.truetype(fp, 20) if fp else ImageFont.load_default()

    start_box = (540, 35, 740, 100)
    result_box = (435, 510, 845, 680)
    end_box = (540, 1010, 740, 1080)
    node_fill = "#F6D1AD"
    line_color = "#2B2B2B"

    draw_node(draw, start_box, "开始", font_node, fill=node_fill, rounded=True)

    raw_sources = list(record.get("sources") or [])
    display_sources = raw_sources[:5]
    if len(raw_sources) > 5:
        display_sources = raw_sources[:4] + [f"其他{len(raw_sources) - 4}张源表"]
    if not display_sources:
        display_sources = ["程序配置源表"]

    if len(display_sources) == 1:
        xs = [490]
        y = 290
        boxes = [(x, y, x + 300, y + 105) for x in xs]
    elif len(display_sources) == 2:
        xs = [245, 735]
        y = 290
        boxes = [(x, y, x + 300, y + 105) for x in xs]
    elif len(display_sources) == 3:
        xs = [34, 449, 864]
        y = 290
        boxes = [(x, y, x + 300, y + 105) for x in xs]
    else:
        xs_top = [34, 344, 654, 964]
        boxes = [(x, 245, x + 280, 345) for x in xs_top[: min(4, len(display_sources))]]
        if len(display_sources) == 5:
            boxes.append((500, 385, 780, 485))

    max_source_bottom = max(box[3] for box in boxes)
    routing_gap = 42
    result_gap = 45
    result_h = result_box[3] - result_box[1]
    result_top = max(result_box[1], max_source_bottom + routing_gap + result_gap)
    result_box = (result_box[0], result_top, result_box[2], result_top + result_h)

    for source, box in zip(display_sources, boxes):
        if source.startswith("其他"):
            label = f"提取处理表数据\n{source}"
        else:
            info = record.get("resource_info", {}).get(source, {})
            resource_name = sanitize_stats_logic_text(info.get("资源名称", ""))
            label = f"提取处理表数据\n{resource_name}\n{source}" if resource_name else f"提取处理表数据\n{source}"
        draw_node(draw, box, label, font_small, fill=node_fill, rounded=False)

    # Template-like branch from start to each source, then converge into result.
    branch_y = 195
    start_mid = ((start_box[0] + start_box[2]) // 2, start_box[3])
    draw.line([start_mid, (start_mid[0], branch_y)], fill=line_color, width=4)
    left_x = min(box[0] + (box[2] - box[0]) // 2 for box in boxes)
    right_x = max(box[0] + (box[2] - box[0]) // 2 for box in boxes)
    draw.line([(min(left_x, start_mid[0]), branch_y), (max(right_x, start_mid[0]), branch_y)], fill=line_color, width=4)
    for box in boxes:
        cx = box[0] + (box[2] - box[0]) // 2
        draw_arrow(draw, (cx, branch_y), (cx, box[1]))

    merge_y = max_source_bottom + routing_gap
    for box in boxes:
        cx = box[0] + (box[2] - box[0]) // 2
        draw.line([(cx, box[3]), (cx, merge_y)], fill=line_color, width=4)
    draw.line([(left_x, merge_y), (right_x, merge_y)], fill=line_color, width=4)
    draw_arrow(draw, ((left_x + right_x) // 2, merge_y), ((result_box[0] + result_box[2]) // 2, result_box[1]))

    result_text = f"统计融合为\n{record['result_cn']}\n{record['result_en']}"
    draw_node(draw, result_box, result_text, font_node, fill=node_fill, rounded=False)
    draw_node(draw, end_box, "结束", font_node, fill=node_fill, rounded=True)

    draw_arrow(draw, ((result_box[0] + result_box[2]) // 2, result_box[3]), ((end_box[0] + end_box[2]) // 2, end_box[1]))

    img.save(output_path)


def build_records(
    ledger_path: Path,
    catalog_path: Path,
    flowchart_dir: Path,
    service_dir: str = DEFAULT_SERVICE_DIR,
) -> tuple[list[dict], dict[str, dict]]:
    ledger_rows = load_ledger_rows(ledger_path, service_dir)
    records: list[dict] = []
    needed_resources: set[str] = set()
    for row in ledger_rows:
        result_cn, result_en = parse_result_name(row.get(LEDGER_RESULT_COL, ""))
        nodes, edges = parse_xml_program(row.get(LEDGER_XML_COL, ""))
        sqls = [node["sql"] for node in nodes if node.get("sql")]
        sources = source_tables(sqls, result_en)
        fields = parse_result_fields_from_sql(sqls, result_en)
        record = {
            "ledger_row": int(row["_row"]),
            "result_cn": result_cn,
            "result_en": result_en,
            "nodes": nodes,
            "edges": edges,
            "sources": sources,
            "fields": fields,
            "logic_steps": [],
        }
        records.append(record)
        needed_resources.add(result_en)
        needed_resources.update(sources)

    resource_info, needed_dir_codes = load_resource_info(catalog_path, needed_resources)
    catalog_fields = load_catalog_fields(catalog_path, needed_dir_codes)

    for record in records:
        if not record["fields"]:
            info = resource_info.get(record["result_en"], {})
            directory_code = str(info.get("数据目录代码", "")).strip()
            record["fields"] = catalog_fields.get(directory_code, [])
        record["field_source"] = record["fields"][0].get("来源", "") if record["fields"] else "未匹配"
        record["resource_info"] = resource_info
        record["logic_steps"] = build_business_logic_steps(record, resource_info)

    flowchart_dir.mkdir(parents=True, exist_ok=True)
    for idx, record in enumerate(records, start=1):
        img_path = flowchart_dir / f"flowchart-{idx:02d}-{record['result_en']}.png"
        draw_flowchart_png(record, img_path)
        record["flowchart"] = str(img_path)
    return records, resource_info


def fill_source_list(ws, records: list[dict], resource_info: dict[str, dict]) -> None:
    clear_sheet(ws)
    headers = ["序号", "资源编目（非必填）", "资源名称", "数据提供方", "资源类型", "资源信息（表名）", "数据融合加工表"]
    add_header_row(ws, headers)
    rows = []
    seen: set[tuple[str, str]] = set()
    for record in records:
        for source in record["sources"]:
            key = (source, record["result_en"])
            if key in seen:
                continue
            seen.add(key)
            info = resource_info.get(source, {})
            rows.append(
                [
                    len(rows) + 1,
                    info.get("数据目录代码", ""),
                    info.get("资源名称", ""),
                    DEFAULT_DATA_PROVIDER,
                    DEFAULT_RESOURCE_TYPE,
                    source,
                    record["result_en"],
                ]
            )
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    style_range(ws, 1, max(1, len(rows) + 1), 1, len(headers))
    set_widths(ws, [8, 22, 30, 28, 12, 34, 40])
    setup_common_ws(ws)


def fill_relation(ws, records: list[dict]) -> None:
    clear_sheet(ws)
    row = 1
    for idx, record in enumerate(records, start=1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        title = ws.cell(row, 1, f"2.{idx} {record['result_en']}")
        title.font = Font(name=DEFAULT_FONT_NAME, size=12, bold=True)
        title.fill = PatternFill("solid", fgColor=DEFAULT_HIGHLIGHT_FILL)
        title.alignment = Alignment(horizontal="left", vertical="center")
        style_range(ws, row, row, 1, 3)

        ws.cell(row + 1, 1, "")
        ws.cell(row + 1, 2, "文字描述")
        ws.cell(row + 1, 3, "\n".join(f"{i + 1}\t{step}" for i, step in enumerate(record["logic_steps"])))
        ws.cell(row + 2, 1, "")
        ws.cell(row + 2, 2, "数据处理流程图")
        img = XLImage(record["flowchart"])
        img.width = 760
        img.height = 665
        ws.add_image(img, f"C{row + 2}")
        ws.row_dimensions[row].height = 28
        ws.row_dimensions[row + 1].height = max(80, min(220, 28 * len(record["logic_steps"])))
        ws.row_dimensions[row + 2].height = 500
        for rr in (row + 1, row + 2):
            ws.cell(rr, 2).font = Font(name=DEFAULT_FONT_NAME, size=11, bold=True)
            ws.cell(rr, 2).fill = PatternFill("solid", fgColor=DEFAULT_HIGHLIGHT_FILL)
        style_range(ws, row + 1, row + 2, 1, 3)
        ws.cell(row + 1, 3).alignment = Alignment(vertical="top", wrap_text=True)
        if idx < len(records):
            ws.row_dimensions[row + 3].height = 20
        row += 4 if idx < len(records) else 3
    set_widths(ws, [6, 20, 120])
    setup_common_ws(ws)


def fill_result_list(ws, records: list[dict], resource_info: dict[str, dict]) -> None:
    clear_sheet(ws)
    headers = ["序号", "资源编目（非必填）", "资源名称", "数据提供方", "资源类型", "数据统计分析表的资源信息（表名）"]
    add_header_row(ws, headers)
    for r_idx, record in enumerate(records, start=2):
        info = resource_info.get(record["result_en"], {})
        row = [
            r_idx - 1,
            info.get("数据目录代码", ""),
            record["result_cn"],
            info.get("单位名称", "") or DEFAULT_DATA_PROVIDER,
            info.get("资源类型", "") or DEFAULT_RESOURCE_TYPE,
            record["result_en"],
        ]
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    style_range(ws, 1, len(records) + 1, 1, len(headers))
    set_widths(ws, [8, 22, 42, 28, 12, 45])
    setup_common_ws(ws)


def fill_result_detail(ws, records: list[dict]) -> None:
    clear_sheet(ws)
    headers = ["", "字段中文名", "字段英文名", "字段类型", "默认", "不可为空", "唯一", "字段注释"]
    row = 1
    for idx, record in enumerate(records, start=1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row, 1, f"4.{idx} {record['result_cn']}")
        cell.font = Font(name=DEFAULT_FONT_NAME, size=12, bold=True)
        cell.fill = PatternFill("solid", fgColor=DEFAULT_HIGHLIGHT_FILL)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        style_range(ws, row, row, 1, 8)

        ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 1, end_column=8)
        ws.cell(row + 1, 2, record["result_en"])
        ws.cell(row + 1, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row + 1, 2).font = Font(name=DEFAULT_FONT_NAME, size=11, bold=True)
        style_range(ws, row + 1, row + 1, 1, 8)

        for c_idx, header in enumerate(headers, start=1):
            c = ws.cell(row + 2, c_idx, header)
            c.font = Font(name=DEFAULT_FONT_NAME, size=11, bold=True)
            c.fill = PatternFill("solid", fgColor=DEFAULT_HIGHLIGHT_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        style_range(ws, row + 2, row + 2, 1, 8)

        fields = record["fields"] or [
            {
                "字段中文名": "未解析到字段信息",
                "字段英文名": "",
                "字段类型": "",
                "默认": "",
                "不可为空": "",
                "唯一": "",
                "字段注释": "请根据程序 XML 或数据目录手动补充",
            }
        ]
        for offset, field in enumerate(fields, start=3):
            rr = row + offset
            values = [
                "",
                field.get("字段中文名", ""),
                field.get("字段英文名", ""),
                field.get("字段类型", ""),
                field.get("默认", ""),
                field.get("不可为空", ""),
                field.get("唯一", ""),
                field.get("字段注释", ""),
            ]
            for c_idx, value in enumerate(values, start=1):
                ws.cell(rr, c_idx, value)
        end_row = row + 2 + len(fields)
        style_range(ws, row, end_row, 1, 8)
        for rr in range(row + 3, end_row + 1):
            ws.cell(rr, 1).fill = PatternFill("solid", fgColor="FFFFFF")
        row = end_row + 2
    set_widths(ws, [6, 24, 28, 18, 12, 12, 10, 36])
    setup_common_ws(ws)


def build_workbook(template_path: Path, output_path: Path, records: list[dict], resource_info: dict[str, dict]) -> None:
    wb = load_workbook(template_path)
    fill_source_list(wb[SHEET_SOURCE_LIST], records, resource_info)
    fill_relation(wb[SHEET_RELATION], records)
    fill_result_list(wb[SHEET_RESULT_LIST], records, resource_info)
    fill_result_detail(wb[SHEET_RESULT_DETAIL], records)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                font = copy(cell.font)
                font.name = DEFAULT_FONT_NAME
                cell.font = font
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_stats_result_usage_workbook(
    ledger_path: str | Path,
    service_dir: str,
    template_path: str | Path,
    output_path: str | Path,
    catalog_path: str | Path,
) -> str:
    ledger = Path(ledger_path)
    template = Path(template_path)
    catalog = Path(catalog_path)
    output = Path(output_path)
    flowchart_dir = output.parent / f"{output.stem}_flowcharts"
    records, resource_info = build_records(ledger, catalog, flowchart_dir, service_dir)
    build_workbook(template, output, records, resource_info)
    return str(output)


def main() -> int:
    if len(sys.argv) == 5:
        ledger = Path(sys.argv[1])
        service_dir = DEFAULT_SERVICE_DIR
        template = Path(sys.argv[2])
        catalog = Path(sys.argv[3])
        output = Path(sys.argv[4])
    elif len(sys.argv) == 6:
        ledger = Path(sys.argv[1])
        service_dir = sys.argv[2]
        template = Path(sys.argv[3])
        catalog = Path(sys.argv[4])
        output = Path(sys.argv[5])
    else:
        print(
            "usage: build_stats_result_usage_workbook.py LEDGER [SERVICE_DIR] TEMPLATE CATALOG OUTPUT",
            file=sys.stderr,
        )
        return 2
    flowchart_dir = output.parent / f"{output.stem}_flowcharts"
    records, resource_info = build_records(ledger, catalog, flowchart_dir, service_dir)
    build_workbook(template, output, records, resource_info)
    print(json.dumps({
        "output": str(output),
        "service_dir": service_dir,
        "records": len(records),
        "source_rows": sum(len(record["sources"]) for record in records),
        "flowcharts": len(records),
        "field_sources": {record["result_en"]: record["field_source"] for record in records},
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
