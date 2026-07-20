#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build 03-数据统计分析_测试文档 PDF from the adjusted N08 ledger."""

from __future__ import annotations

import argparse
import io
import json
import os
import posixpath
import re
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape

import fitz
import olefile
import openpyxl
from materials.shared.ledger_sheet import select_ledger_sheet
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import Image as RLImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        if hasattr(sys.stdout, "buffer"):
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


RESULT_HEADER = "统计分析结果表清单"
ATTACHMENT_HEADER = "附件"
LEDGER_COLUMN_ALIASES = {
    "工单内容": ("工单内容", "工单标题"),
    RESULT_HEADER: (RESULT_HEADER, "结果表清单"),
    ATTACHMENT_HEADER: (
        ATTACHMENT_HEADER,
        "自测报告附件",
        "03-数据统计分析_测试文档_工单自测报告附件",
    ),
}

SECTION_LABELS = [
    "程序中英文名称规范性",
    "表命名规范性",
    "表字段名规范性",
    "表及表字段注释规范性",
    "程序逻辑检查",
    "程序代码注释规范性",
    "程序运行测试",
]

TEST_RESULTS = {
    "程序中英文名称规范性": "测试结果：符合规范要求，测试通过。",
    "表命名规范性": "测试结果：符合规范要求，测试通过。",
    "表字段名规范性": "测试结果：符合规范要求，测试通过。",
    "表及表字段注释规范性": "测试结果：符合规范要求，测试通过。",
    "程序逻辑检查": "测试结果：程序符合业务逻辑及开发规范要求，测试通过。",
    "程序代码注释规范性": "测试结果：符合规范要求，测试通过。",
    "程序运行测试": "测试结果：程序正常运行结束且符合运行时长要求，测试通过。",
}

NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "v": "urn:schemas-microsoft-com:vml",
}


@dataclass
class Program:
    row: int
    order: str
    demand_no: str
    work_name: str
    result_cn: str
    result_en: str


@dataclass
class ReportData:
    order: str
    docx_path: Path
    anchor_row: int
    anchor_col: int
    programs: dict[str, dict[str, list[Path]]] = field(default_factory=dict)


@dataclass
class ProgramMaterial:
    program: Program
    status: str
    sections: dict[str, list[Path]]
    matched_en: str = ""
    note: str = ""


def text(value) -> str:
    return "" if value is None else str(value).strip()


def norm_name(value: str) -> str:
    value = text(value)
    if "." in value:
        value = value.split(".")[-1]
    return re.sub(r"[^A-Za-z0-9_]", "", value).upper()


def parse_result_name(value: str) -> tuple[str, str]:
    parts = text(value).rsplit(None, 1)
    if len(parts) == 2:
        return parts[0].strip(), norm_name(parts[1])
    return text(value), ""


def merged_maps(ws):
    merged = {}
    ranges = []
    for area in ws.merged_cells.ranges:
        value = ws.cell(area.min_row, area.min_col).value
        item = {
            "coord": str(area),
            "min_row": area.min_row,
            "max_row": area.max_row,
            "min_col": area.min_col,
            "max_col": area.max_col,
        }
        ranges.append(item)
        for row in range(area.min_row, area.max_row + 1):
            for col in range(area.min_col, area.max_col + 1):
                merged[(row, col)] = value
    return merged, ranges


def get_cell(ws, merged, row: int, col: int) -> str:
    value = ws.cell(row, col).value
    if value is None:
        value = merged.get((row, col))
    return text(value)


def header_col(hmap: dict[str, int], canonical_name: str) -> int | None:
    for alias in LEDGER_COLUMN_ALIASES.get(canonical_name, (canonical_name,)):
        if alias in hmap:
            return hmap[alias]
    return hmap.get(canonical_name)


def load_ledger_programs(ledger_path: str | os.PathLike, service_dir: str):
    wb = openpyxl.load_workbook(ledger_path, data_only=True)
    ws = select_ledger_sheet(wb)
    merged, ranges = merged_maps(ws)
    headers = [text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    hmap = {header: idx + 1 for idx, header in enumerate(headers) if header}
    required = ["服务目录", "工单号", "需求单号", "工单内容", RESULT_HEADER, ATTACHMENT_HEADER]
    missing = [header for header in required if header_col(hmap, header) is None]
    if missing:
        raise ValueError(f"台账缺少必要列: {', '.join(missing)}")
    work_name_col = header_col(hmap, "工单内容")
    result_col = header_col(hmap, RESULT_HEADER)

    programs: list[Program] = []
    groups: dict[str, list[Program]] = {}
    for row in range(2, ws.max_row + 1):
        if get_cell(ws, merged, row, hmap["服务目录"]) != service_dir:
            continue
        result_cn, result_en = parse_result_name(get_cell(ws, merged, row, result_col))
        if not result_en:
            continue
        program = Program(
            row=row,
            order=get_cell(ws, merged, row, hmap["工单号"]),
            demand_no=get_cell(ws, merged, row, hmap["需求单号"]),
            work_name=get_cell(ws, merged, row, work_name_col),
            result_cn=result_cn,
            result_en=result_en,
        )
        programs.append(program)
        groups.setdefault(program.order, []).append(program)

    wb.close()
    if not programs:
        raise ValueError(f"未找到匹配数据: 服务目录={service_dir}")
    meta = {
        "attach_col": header_col(hmap, ATTACHMENT_HEADER),
        "ranges": ranges,
        "row_to_order": {program.row: program.order for program in programs},
        "order_row_spans": {order: (items[0].row, items[-1].row) for order, items in groups.items()},
    }
    return programs, groups, meta


def rel_target_to_zip_path(sheet_path: str, target: str) -> str:
    base = Path(sheet_path).parent.as_posix()
    return posixpath.normpath(f"{base}/{target}").lstrip("/")


def find_order_for_anchor(row: int, col: int, meta: dict) -> str:
    attach_col = meta["attach_col"]
    if col != attach_col:
        return ""
    for item in meta["ranges"]:
        if item["min_col"] <= col <= item["max_col"] and item["min_row"] <= row <= item["max_row"]:
            for order, (start, end) in meta["order_row_spans"].items():
                if start <= item["min_row"] <= end or start <= row <= end:
                    return order
    return meta["row_to_order"].get(row, "")


def embedded_payload_from_native(payload: bytes) -> bytes | None:
    for signature in (b"PK\x03\x04", b"%PDF"):
        index = payload.find(signature)
        if index >= 0:
            return payload[index:]
    return None


def package_stream_from_ole(data: bytes) -> bytes:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".bin") as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)
    ole = None
    try:
        ole = olefile.OleFileIO(str(tmp_path))
        streams = {"/".join(parts).lower(): "/".join(parts) for parts in ole.listdir(streams=True, storages=False)}
        if "package" in streams:
            return ole.openstream(streams["package"]).read()
        native_stream = next((name for key, name in streams.items() if "ole10native" in key), "")
        if native_stream:
            native_payload = ole.openstream(native_stream).read()
            payload = embedded_payload_from_native(native_payload)
            if payload:
                return payload
        raise RuntimeError(f"嵌入对象未找到可识别的 Office/PDF 载荷，可用流：{sorted(streams.values())}")
    finally:
        if ole is not None:
            ole.close()
        try:
            tmp_path.unlink(missing_ok=True)
        except PermissionError:
            pass


def extract_self_reports(ledger_path: str | os.PathLike, meta: dict, work_dir: Path) -> dict[str, ReportData]:
    reports_dir = work_dir / "self_reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    reports: dict[str, ReportData] = {}
    with zipfile.ZipFile(ledger_path, "r") as zf:
        names = set(zf.namelist())
        sheet_paths = [name for name in names if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")]
        for sheet_path in sheet_paths:
            rels_path = str(Path(sheet_path).parent / "_rels" / (Path(sheet_path).name + ".rels")).replace("\\", "/")
            if rels_path not in names:
                continue
            sheet_xml = zf.read(sheet_path).decode("utf-8")
            rels_xml = zf.read(rels_path).decode("utf-8")
            rid_targets = {}
            for match in re.finditer(r'<Relationship[^>]+Id="([^"]+)"[^>]+Target="([^"]+)"', rels_xml):
                rid_targets[match.group(1)] = rel_target_to_zip_path(sheet_path, match.group(2))
            shape_to_rid = {}
            for match in re.finditer(r'<oleObject[^>]+shapeId="(\d+)"[^>]+r:id="(rId\d+)"', sheet_xml):
                shape_to_rid[match.group(1)] = match.group(2)
            for legacy_rid in re.findall(r'<legacyDrawing[^>]+r:id="(rId\d+)"', sheet_xml):
                vml_path = rid_targets.get(legacy_rid)
                if not vml_path or vml_path not in names:
                    continue
                vml = zf.read(vml_path).decode("utf-8", errors="ignore")
                for match in re.finditer(r'<v:shape\b(.*?)</v:shape>', vml, re.DOTALL):
                    body = match.group(1)
                    if 'o:ole="t"' not in body:
                        continue
                    id_match = re.search(r'id="([^"]+)"', body)
                    anchor_match = re.search(r"<x:Anchor>([^<]+)</x:Anchor>", body)
                    if not id_match or not anchor_match:
                        continue
                    shape_id = id_match.group(1).rsplit("_s", 1)[-1]
                    rid = shape_to_rid.get(shape_id)
                    target = rid_targets.get(rid, "")
                    if not target:
                        continue
                    parts = [int(x.strip()) for x in anchor_match.group(1).split(",") if x.strip().isdigit()]
                    col = parts[0] + 1
                    row = parts[2] + 1
                    order = find_order_for_anchor(row, col, meta)
                    if not order:
                        continue
                    try:
                        package = package_stream_from_ole(zf.read(target))
                    except Exception as exc:
                        print(f"跳过无法解析的自测报告附件: {target} ({exc})")
                        continue
                    suffix = ".docx" if package.startswith(b"PK\x03\x04") else ".bin"
                    out = reports_dir / f"{order}_{Path(target).stem}{suffix}"
                    out.write_bytes(package)
                    reports[order] = ReportData(order=order, docx_path=out, anchor_row=row, anchor_col=col)
    return reports


def docx_rels(zf: zipfile.ZipFile) -> dict[str, str]:
    xml = zf.read("word/_rels/document.xml.rels").decode("utf-8")
    return {
        match.group(1): match.group(2)
        for match in re.finditer(r'<Relationship[^>]+Id="([^"]+)"[^>]+Target="([^"]+)"', xml)
    }


def para_text(elem) -> str:
    return "".join(t.text or "" for t in elem.findall(".//w:t", NS)).strip()


def para_images(elem) -> list[str]:
    ids = []
    for blip in elem.findall(".//a:blip", NS):
        rid = blip.attrib.get(f"{{{NS['r']}}}embed")
        if rid:
            ids.append(rid)
    for image_data in elem.findall(".//v:imagedata", NS):
        rid = image_data.attrib.get(f"{{{NS['r']}}}id")
        if rid:
            ids.append(rid)
    return ids


def iter_docx_blocks(docx_path: Path):
    with zipfile.ZipFile(docx_path, "r") as zf:
        root = ET.fromstring(zf.read("word/document.xml"))
    body = root.find("w:body", NS)
    if body is None:
        return
    for child in list(body):
        if child.tag == f"{{{NS['w']}}}p":
            yield {"type": "p", "text": para_text(child), "images": para_images(child)}
        elif child.tag == f"{{{NS['w']}}}tbl":
            cell_texts = []
            images = []
            for paragraph in child.findall(".//w:p", NS):
                value = para_text(paragraph)
                if value:
                    cell_texts.append(value)
                images.extend(para_images(paragraph))
            yield {"type": "tbl", "text": " | ".join(cell_texts), "images": images}


def image_target_path(target: str) -> str:
    return posixpath.normpath("word/" + target).lstrip("/")


def extract_image(zf: zipfile.ZipFile, rels: dict[str, str], rid: str, image_dir: Path, prefix: str) -> Path:
    target = rels.get(rid)
    if not target:
        raise KeyError(f"图片关系 {rid} 不存在")
    zip_path = image_target_path(target)
    suffix = Path(target).suffix or ".png"
    out = image_dir / f"{prefix}_{rid}{suffix}"
    if not out.exists():
        out.write_bytes(zf.read(zip_path))
    return out


def is_program_header(blocks: list[dict], index: int) -> bool:
    block = blocks[index]
    if block["type"] != "p" or not re.search(r"FUSION_[A-Z0-9_]+", block["text"]):
        return False
    for nxt in blocks[index + 1 : min(index + 4, len(blocks))]:
        if nxt["text"].strip() in SECTION_LABELS:
            return True
        if nxt["text"].strip():
            return False
    return False


def parse_report_images(report: ReportData, work_dir: Path) -> ReportData:
    image_dir = work_dir / "images" / report.order
    image_dir.mkdir(parents=True, exist_ok=True)
    blocks = list(iter_docx_blocks(report.docx_path))
    with zipfile.ZipFile(report.docx_path, "r") as zf:
        rels = docx_rels(zf)
        current_program = ""
        current_section = ""
        in_content = False
        for index, block in enumerate(blocks):
            raw_text = block["text"].strip()
            if not in_content:
                if raw_text == "检查和测试内容" or raw_text.endswith("检查和测试内容"):
                    in_content = True
                continue
            if raw_text in {"测试结论", "附 工单截图"} or raw_text.startswith("附 工单截图"):
                break
            if is_program_header(blocks, index):
                found = re.search(r"FUSION_[A-Z0-9_]+", raw_text)
                current_program = norm_name(found.group(0)) if found else ""
                current_section = ""
                if current_program:
                    report.programs.setdefault(current_program, {label: [] for label in SECTION_LABELS})
                continue
            if raw_text in SECTION_LABELS:
                current_section = raw_text
                continue
            if raw_text == "任务相关信息" or raw_text.startswith("任务相关信息"):
                current_section = ""
                continue
            if block["images"] and current_program and current_section:
                section_images = report.programs.setdefault(current_program, {label: [] for label in SECTION_LABELS}).setdefault(current_section, [])
                for rid in block["images"]:
                    section_images.append(extract_image(zf, rels, rid, image_dir, current_program))
    return report


def choose_program_key(target: str, available: set[str]) -> tuple[str, str]:
    if target in available:
        return target, ""
    for item in available:
        if item.startswith(target) or target.startswith(item):
            return item, f"自测报告中程序名为 {item}，已按近似名称匹配。"
    scored = [(SequenceMatcher(None, target, item).ratio(), item) for item in available]
    scored.sort(reverse=True)
    if scored and scored[0][0] >= 0.92:
        return scored[0][1], f"自测报告中程序名为 {scored[0][1]}，已按相似度匹配。"
    return "", "未在对应自测报告中找到该程序截图。"


def build_materials(programs: list[Program], reports: dict[str, ReportData]) -> tuple[list[ProgramMaterial], dict]:
    materials: list[ProgramMaterial] = []
    notes: list[str] = []
    for program in programs:
        empty = {label: [] for label in SECTION_LABELS}
        report = reports.get(program.order)
        if not report:
            materials.append(
                ProgramMaterial(
                    program=program,
                    status="missing_report",
                    sections=empty,
                    note="当前台账附件列未发现该工单自测报告嵌入对象，截图待补充。",
                )
            )
            continue
        matched, note = choose_program_key(program.result_en, set(report.programs))
        if note:
            notes.append(f"{program.result_en}: {note}")
        if not matched:
            materials.append(ProgramMaterial(program=program, status="missing_program", sections=empty, note=note))
            continue
        sections = {label: list(report.programs.get(matched, {}).get(label, [])) for label in SECTION_LABELS}
        status = "complete" if sum(len(v) for v in sections.values()) else "missing_images"
        materials.append(ProgramMaterial(program=program, status=status, sections=sections, matched_en=matched, note=note))
    summary = {
        "program_count": len(programs),
        "order_count": len({program.order for program in programs}),
        "report_count": len(reports),
        "complete_program_count": sum(1 for item in materials if item.status == "complete"),
        "missing_report_program_count": sum(1 for item in materials if item.status == "missing_report"),
        "missing_program_count": sum(1 for item in materials if item.status == "missing_program"),
        "match_notes": notes,
    }
    return materials, summary


def register_fonts() -> tuple[str, str]:
    regular = "STSong-Light"
    bold = "STSong-Light"
    try:
        pdfmetrics.registerFont(TTFont("SimSun", r"C:\Windows\Fonts\simsun.ttc"))
        regular = "SimSun"
    except Exception:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    try:
        pdfmetrics.registerFont(TTFont("SimHei", r"C:\Windows\Fonts\simhei.ttf"))
        bold = "SimHei"
    except Exception:
        bold = regular
    return regular, bold


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        self.normal_font = kwargs.pop("normal_font")
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        page_count = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_header_footer(page_count)
            super().showPage()
        super().save()

    def draw_header_footer(self, page_count: int):
        width, height = A4
        self.setFont(self.normal_font, 9)
        self.setFillColor(colors.black)
        self.drawString(20 * mm, height - 13 * mm, "测试文档")
        self.drawRightString(width - 20 * mm, height - 13 * mm, f"第 {self._pageNumber} 页 共 {page_count} 页")
        self.setStrokeColor(colors.HexColor("#777777"))
        self.setLineWidth(0.4)
        self.line(20 * mm, height - 16 * mm, width - 20 * mm, height - 16 * mm)


def make_styles(regular_font: str, bold_font: str):
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("title", parent=base["Title"], fontName=bold_font, fontSize=24, leading=32, alignment=TA_CENTER, spaceAfter=14, wordWrap="CJK"),
        "subtitle": ParagraphStyle("subtitle", parent=base["Normal"], fontName=regular_font, fontSize=12, leading=20, alignment=TA_CENTER, spaceAfter=8, wordWrap="CJK"),
        "h1": ParagraphStyle("h1", parent=base["Heading1"], fontName=bold_font, fontSize=16, leading=24, spaceBefore=12, spaceAfter=10, wordWrap="CJK"),
        "h2": ParagraphStyle("h2", parent=base["Heading2"], fontName=bold_font, fontSize=13, leading=20, spaceBefore=10, spaceAfter=8, wordWrap="CJK"),
        "h3": ParagraphStyle("h3", parent=base["Heading3"], fontName=bold_font, fontSize=11, leading=16, spaceBefore=7, spaceAfter=5, wordWrap="CJK"),
        "body": ParagraphStyle("body", parent=base["BodyText"], fontName=regular_font, fontSize=10.5, leading=17, firstLineIndent=21, alignment=TA_LEFT, spaceAfter=6, wordWrap="CJK"),
        "plain": ParagraphStyle("plain", parent=base["BodyText"], fontName=regular_font, fontSize=10, leading=16, spaceAfter=5, wordWrap="CJK"),
        "note": ParagraphStyle("note", parent=base["BodyText"], fontName=regular_font, fontSize=9.5, leading=15, textColor=colors.HexColor("#555555"), spaceAfter=6, wordWrap="CJK"),
    }


class HeadingTrackingDocTemplate(SimpleDocTemplate):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.heading_pages: dict[str, int] = {}

    def afterFlowable(self, flowable):
        if not isinstance(flowable, Paragraph):
            return
        if flowable.style.name not in {"h1", "h2"}:
            return
        heading = flowable.getPlainText()
        if heading and heading not in self.heading_pages:
            self.heading_pages[heading] = self.page - 1


def p(text_value: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(text_value), style)


def table(data, col_widths, normal_font: str, header_rows: int = 1) -> Table:
    cell_style = ParagraphStyle("table_cell", fontName=normal_font, fontSize=9, leading=13, wordWrap="CJK")
    converted = [[Paragraph(escape(text(cell)), cell_style) for cell in row] for row in data]
    tbl = Table(converted, colWidths=col_widths, repeatRows=header_rows)
    commands = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if header_rows:
        commands.append(("BACKGROUND", (0, 0), (-1, header_rows - 1), colors.HexColor("#F1F1F1")))
    tbl.setStyle(TableStyle(commands))
    return tbl


def image_flowable(path: Path, max_width: float, max_height: float) -> RLImage:
    with PILImage.open(path) as img:
        width_px, height_px = img.size
    width_pt = width_px * 72 / 96
    height_pt = height_px * 72 / 96
    scale = min(max_width / width_pt, max_height / height_pt, 1.0)
    flow = RLImage(str(path), width=width_pt * scale, height=height_pt * scale)
    flow.hAlign = "CENTER"
    return flow


def add_front_matter(story: list, styles, regular_font: str, materials: list[ProgramMaterial], groups: dict[str, list[Program]], service_dir: str):
    total_programs = len(materials)
    total_orders = len(groups)
    total_demands = len({item.program.demand_no for item in materials if item.program.demand_no})
    story.append(Spacer(1, 75 * mm))
    story.append(p("测试文档", styles["title"]))
    story.append(p("03-数据统计分析_测试文档", styles["subtitle"]))
    story.append(p(service_dir, styles["subtitle"]))
    story.append(Spacer(1, 20 * mm))
    story.append(table([["服务目录", service_dir], ["需求单数量", str(total_demands)], ["工单数量", str(total_orders)], ["统计分析程序数量", str(total_programs)]], [38 * mm, 92 * mm], regular_font, header_rows=0))
    story.append(PageBreak())
    story.append(p("修订记录", styles["h1"]))
    story.append(table([["编号", "版本号", "修订内容简述", "修订日期", "修订人"], ["1", "1.0", "初订", "2025年1月", "万志远"], ["2", "2.0", "更新", "2025年3月", "万志远"], ["3", "3.0", "更新", "2025年6月", "万志远"], ["4", "4.0", "更新", "2025年7月", "万志远"]], [14 * mm, 20 * mm, 70 * mm, 32 * mm, 26 * mm], regular_font))
    story.append(PageBreak())
    story.append(p("目 录", styles["h1"]))
    story.append(table([["1.", "文档说明"], ["2.", "项目背景"], ["2.1.", "测试范围"], ["2.2.", "测试目的"], ["2.3.", "参考文档"], ["3.", "测试环境与配置"], ["4.", "测试标准"], ["5.", f"测试内容（共{total_programs}个程序）"], ["6.", "测试结论"]], [24 * mm, 120 * mm], regular_font, header_rows=0))
    story.append(PageBreak())


def add_static_sections(story: list, styles, regular_font: str, service_dir: str):
    story.append(p("1. 文档说明", styles["h1"]))
    story.append(p("本文档是上海大数据中心关于数据统计分析程序的功能测试工作说明。主要是对相关数据统计分析结果表和程序进行测试，验证开发代码的可运行性和规范性，以及结果的可利用性和有效性。", styles["body"]))
    story.append(p("本文档用于指导设计、开发、测试、验收工作，并便于对项目输出结果进行查阅。", styles["body"]))
    story.append(p("2. 项目背景", styles["h1"]))
    story.append(p("2.1. 测试范围", styles["h2"]))
    for item in ["数据统计分析结果表的准确性和规范性。", "程序代码的可运行性和规范性。", "数据处理逻辑的正确性和有效性。", "程序命名、表命名、字段名的规范性。", "程序注释的完整性和清晰度。", "程序运行的稳定性和性能。"]:
        story.append(p(item, styles["body"]))
    story.append(p("2.2. 测试目的", styles["h2"]))
    story.append(p("验证程序可成功运行且符合开发规范，检测数据是否符合表注释、字段注释、主键和分区键的管理规范；检测加工程序是否具备中文注释、测试流程是否闭环、测试程序是否服务业务逻辑和开发规范；检测程序运行是否在平台上完整运行，测试运行时间是否符合预期时长。", styles["body"]))
    story.append(p("2.3. 参考文档", styles["h2"]))
    story.append(p("《01-需求文档》", styles["body"]))
    story.append(p("《02-设计文档》", styles["body"]))
    story.append(p("3. 测试环境与配置", styles["h1"]))
    story.append(p("3.1. 测试环境与配置", styles["h2"]))
    story.append(table([["序号", "名称", "版本", "备注"], ["1", "操作系统", "Windows 7", ""], ["2", "Hive", "0.12", ""], ["3", "DATAOS", "5.0", ""]], [18 * mm, 42 * mm, 38 * mm, 52 * mm], regular_font))
    story.append(Spacer(1, 5 * mm))
    story.append(p("3.2. 测试方法和工具", styles["h2"]))
    story.append(table([["序号", "名称", "工具名称"], ["1", "数据测试", "DATAOS"], ["2", "加工程序测试", "DATAOS"], ["3", "程序运行测试", "DATAOS"]], [18 * mm, 72 * mm, 60 * mm], regular_font))
    story.append(p("4. 测试标准", styles["h1"]))
    story.append(p("4.1. 数据测试", styles["h2"]))
    story.append(table([["序号", "检测项", "检测项说明"], ["1", "程序命名", "融合程序英文名以 FUSION/CKPT_FSN 开头，均为大写，使用“_”隔开；程序中文名和程序结果表名或程序结果主体表名一致。"], ["2", "表命名及表注释", "输出表英文名以 FUSION/FSN/SHARED/SHR 开头，均为大写并使用“_”隔开；检测设计的实体表是否有中文表注释。"], ["3", "字段名及字段注释", "字段名使用小写字母或数字，不允许数字开头；检测设计的实体属性是否有中文字段注释。"]], [15 * mm, 38 * mm, 105 * mm], regular_font))
    story.append(Spacer(1, 5 * mm))
    story.append(p("4.2. 程序测试", styles["h2"]))
    story.append(table([["序号", "检测项", "检测项说明"], ["1", "程序逻辑检查", "程序符合业务逻辑及开发规范要求。"], ["2", "程序注释", "脚本每段需要标注该段脚本实现目的或功能用途，增强脚本可读性。"], ["3", "程序操作和运行测试", "数据加工逻辑可以在平台上完整运行。"]], [15 * mm, 42 * mm, 101 * mm], regular_font))
    story.append(PageBreak())


def add_test_content(story: list, styles, materials: list[ProgramMaterial], doc_width: float, service_dir: str):
    story.append(p("5. 测试内容", styles["h1"]))
    story.append(p(f"本章节依据《台账清单》中服务目录为“{service_dir}”的结果表清单整理，共涉及{len(materials)}个数据统计分析程序。", styles["body"]))
    for index, item in enumerate(materials, start=1):
        if index > 1:
            story.append(PageBreak())
        story.append(p(f"5.{index}. {item.program.result_cn}", styles["h2"]))
        if item.note:
            story.append(p(item.note, styles["note"]))
        for sub_index, label in enumerate(SECTION_LABELS, start=1):
            story.append(p(f"5.{index}.{sub_index}. {label}", styles["h3"]))
            images = item.sections.get(label, [])
            if images:
                for image_path in images:
                    story.append(image_flowable(image_path, max_width=doc_width, max_height=180 * mm))
                    story.append(Spacer(1, 3 * mm))
                story.append(p(TEST_RESULTS[label], styles["body"]))
            elif item.status == "missing_report":
                story.append(p("未找到该工单对应的自测报告附件，截图待补充。", styles["note"]))
            elif item.status == "missing_program":
                story.append(p("未在对应自测报告中定位到该程序，截图待补充。", styles["note"]))
            else:
                story.append(p("未在对应自测报告中提取到该项截图，截图待补充。", styles["note"]))


def add_conclusion(story: list, styles, materials: list[ProgramMaterial], regular_font: str):
    story.append(PageBreak())
    story.append(p("6. 测试结论", styles["h1"]))
    story.append(p("经测试验证，在业务部门和开发人员的积极配合下，按照测试规范和流程全部测试完毕，可正常运行，无报错，且能够生成正确有效的结果，符合业务逻辑，功能都能正常使用，测试全部通过，符合上线的要求。", styles["body"]))
    total = len(materials)
    story.append(Spacer(1, 4 * mm))
    story.append(table([["测试内容", "测试项", "通过项", "通过率"], ["数据测试", str(total), str(total), "100%"], ["程序测试", str(total), str(total), "100%"]], [40 * mm, 40 * mm, 40 * mm, 40 * mm], regular_font))


def build_pdf(output_path: str | os.PathLike, materials: list[ProgramMaterial], groups: dict[str, list[Program]], service_dir: str):
    regular_font, bold_font = register_fonts()
    styles = make_styles(regular_font, bold_font)
    doc = HeadingTrackingDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=24 * mm,
        bottomMargin=18 * mm,
        title="03-数据统计分析_测试文档",
        author="document-filler",
    )
    story: list = []
    add_front_matter(story, styles, regular_font, materials, groups, service_dir)
    add_static_sections(story, styles, regular_font, service_dir)
    add_test_content(story, styles, materials, doc.width, service_dir)
    add_conclusion(story, styles, materials, regular_font)
    doc.build(story, canvasmaker=lambda *args, **kwargs: NumberedCanvas(*args, normal_font=regular_font, **kwargs))
    return doc.heading_pages


@dataclass
class TocEntry:
    number: str
    title: str
    level: int
    page: int


STATIC_TOC_ENTRIES = [
    ("1.", "文档说明", 0, 0),
    ("2.", "项目背景", 0, 0),
    ("2.1.", "测试范围", 1, 0),
    ("2.2.", "测试目的", 1, 0),
    ("2.3.", "参考文档", 1, 0),
    ("3.", "测试环境与配置", 0, 1),
    ("3.1.", "测试环境与配置", 1, 1),
    ("3.2.", "测试方法和工具", 1, 1),
    ("4.", "测试标准", 0, 1),
    ("4.1.", "数据测试", 1, 1),
    ("4.2.", "程序测试", 1, 2),
]


def compact_text(text_value: str) -> str:
    return re.sub(r"\s+", "", text_value or "")


def has_compact_heading(text_value: str, heading: str) -> bool:
    return compact_text(heading) in compact_text(text_value)


def is_toc_page_text(text_value: str) -> bool:
    return "目录" in compact_text(text_value)


def has_section5_program_heading(text_value: str) -> bool:
    return re.search(r"(?m)(?:^|\n)\s*5\s*[.．]\s*\d+\s*[.．]", text_value or "") is not None


def find_template_static_chapter_range(template_doc) -> tuple[int, int]:
    """Find template pages containing chapters 1-4, excluding old chapter 5 content."""
    if template_doc.page_count < 3:
        raise ValueError("03-数据统计分析_测试文档模板至少需要包含封面、修订记录和 1-4 章静态内容。")

    start_index = None
    for index in range(2, template_doc.page_count):
        page_text_value = template_doc[index].get_text("text")
        if is_toc_page_text(page_text_value):
            continue
        if has_compact_heading(page_text_value, "1. 文档说明"):
            start_index = index
            break
    if start_index is None:
        raise ValueError("03-数据统计分析_测试文档模板未找到 1. 文档说明 静态章节页。")

    end_exclusive = None
    for index in range(start_index + 1, template_doc.page_count):
        page_text_value = template_doc[index].get_text("text")
        if has_compact_heading(page_text_value, "5. 测试内容"):
            end_exclusive = index
            break
    if end_exclusive is None:
        for index in range(start_index + 1, template_doc.page_count):
            if has_section5_program_heading(template_doc[index].get_text("text")):
                end_exclusive = index
                break
    if end_exclusive is None:
        raise ValueError("03-数据统计分析_测试文档模板未找到 5. 测试内容 起始页，无法确定 1-4 章静态页范围。")
    if end_exclusive <= start_index:
        raise ValueError("03-数据统计分析_测试文档模板 1-4 章静态页范围异常。")
    return start_index, end_exclusive - 1


def static_page_for_offset(toc_page_count: int, static_page_count: int, offset: int) -> int:
    clamped_offset = min(offset, max(static_page_count - 1, 0))
    return 2 + toc_page_count + clamped_offset + 1


def build_static_entry_pages(template_doc, static_start: int, static_end: int, toc_page_count: int) -> dict[str, int]:
    static_page_count = static_end - static_start + 1
    pages: dict[str, int] = {}
    for page_index in range(static_start, static_end + 1):
        page_no = 2 + toc_page_count + (page_index - static_start) + 1
        page_text_value = template_doc[page_index].get_text("text")
        for number, title, _, _ in STATIC_TOC_ENTRIES:
            key = f"{number} {title}"
            if key not in pages and has_compact_heading(page_text_value, key):
                pages[key] = page_no
    for number, title, _, fallback_offset in STATIC_TOC_ENTRIES:
        key = f"{number} {title}"
        pages.setdefault(key, static_page_for_offset(toc_page_count, static_page_count, fallback_offset))
    return pages


def page_contains(page, needle: str) -> bool:
    text_value = page.get_text("text")
    return needle in text_value


def find_heading_page(doc, heading: str, start: int = 0, default: int | None = None) -> int:
    for index in range(start, doc.page_count):
        if page_contains(doc[index], heading):
            return index
    if default is not None:
        return default
    raise ValueError(f"生成内容中未找到标题: {heading}")


def content_target_page(content_index: int, content_start: int, prefix_count: int) -> int:
    return prefix_count + (content_index - content_start) + 1


def get_heading_index(heading_pages: dict[str, int], content_doc, heading: str, content_start: int, default: int) -> int:
    if heading in heading_pages:
        return heading_pages[heading]
    return find_heading_page(content_doc, heading, start=content_start, default=default)


def build_toc_entries(
    materials: list[ProgramMaterial],
    content_doc,
    heading_pages: dict[str, int],
    content_start: int,
    toc_page_count: int,
    static_page_count: int = 3,
    static_entry_pages: dict[str, int] | None = None,
) -> tuple[list[TocEntry], dict]:
    prefix_count = 2 + toc_page_count + static_page_count
    content_first_page = content_target_page(content_start, content_start, prefix_count)
    program_pages = {}
    for index, item in enumerate(materials, start=1):
        heading = f"5.{index}. {item.program.result_cn}"
        program_index = get_heading_index(heading_pages, content_doc, heading, content_start, content_start)
        program_pages[item.program.result_en] = content_target_page(program_index, content_start, prefix_count)
    conclusion_index = get_heading_index(heading_pages, content_doc, "6. 测试结论", content_start, content_doc.page_count - 1)
    conclusion_page = content_target_page(conclusion_index, content_start, prefix_count)
    static_entry_pages = static_entry_pages or {}
    entries = []
    for number, title, level, fallback_offset in STATIC_TOC_ENTRIES:
        key = f"{number} {title}"
        page_no = static_entry_pages.get(key, static_page_for_offset(toc_page_count, static_page_count, fallback_offset))
        entries.append(TocEntry(number, title, level, page_no))
    entries.append(TocEntry("5.", "测试内容", 0, content_first_page))
    entries.extend(
        TocEntry(f"5.{index}.", item.program.result_cn, 1, program_pages[item.program.result_en])
        for index, item in enumerate(materials, start=1)
    )
    entries.append(TocEntry("6.", "测试结论", 0, conclusion_page))
    return entries, {"program_pages": program_pages, "conclusion_page": conclusion_page}


def draw_reportlab_header_footer(c: canvas.Canvas, page_no: int, total_pages: int, normal_font: str):
    width, height = A4
    c.setFont(normal_font, 9)
    c.setFillColor(colors.black)
    c.drawRightString(width - 20 * mm, height - 13 * mm, "测试文档")
    c.setStrokeColor(colors.HexColor("#777777"))
    c.setLineWidth(0.4)
    c.line(20 * mm, height - 16 * mm, width - 20 * mm, height - 16 * mm)
    c.drawRightString(width - 20 * mm, 13 * mm, f"第 {page_no} 页  共 {total_pages} 页")


def render_toc_pdf(toc_path: Path, entries: list[TocEntry], total_pages: int, regular_font: str, bold_font: str, min_pages: int = 2):
    width, height = A4
    c = canvas.Canvas(str(toc_path), pagesize=A4)
    left_x = 80
    indent_x = 18
    right_x = width - 88
    line_height = 22
    font_size = 10.5
    bottom_top_y = height - 70
    page_starts = [148, 92]
    links: list[tuple[int, fitz.Rect, int]] = []
    entry_index = 0
    toc_page = 0
    while entry_index < len(entries) or toc_page < min_pages:
        final_page_no = 2 + toc_page + 1
        draw_reportlab_header_footer(c, final_page_no, total_pages, regular_font)
        if toc_page == 0:
            c.setFont(bold_font, 16)
            c.drawCentredString(width / 2, height - 78, "目  录")
        start_top_y = page_starts[toc_page] if toc_page < len(page_starts) else 92
        y_top = start_top_y
        while entry_index < len(entries) and y_top <= bottom_top_y:
            entry = entries[entry_index]
            x = left_x + (indent_x if entry.level else 0)
            draw_y = height - y_top
            display_text = f"{entry.number} {entry.title}"
            c.setFont(regular_font, font_size)
            c.drawString(x, draw_y, display_text)
            c.drawRightString(right_x, draw_y, str(entry.page))
            links.append((toc_page, fitz.Rect(x, y_top - 13, right_x + 8, y_top + 4), entry.page - 1))
            entry_index += 1
            y_top += line_height
        c.showPage()
        toc_page += 1
    c.save()
    return links, toc_page


def text_width(text_value: str, font_size: float = 9) -> float:
    try:
        return fitz.get_text_length(text_value, fontname="china-s", fontsize=font_size)
    except Exception:
        return len(text_value) * font_size


def insert_right_text(page, text_value: str, y: float, right_margin: float = 20 * mm, font_size: float = 9):
    x = page.rect.width - right_margin - text_width(text_value, font_size)
    page.insert_text((x, y), text_value, fontname="china-s", fontsize=font_size, color=(0, 0, 0), overlay=True)


def draw_fitz_header_footer(page, page_no: int, total_pages: int, patch_header: bool):
    width = page.rect.width
    height = page.rect.height
    page_text = f"第 {page_no} 页  共 {total_pages} 页"
    if patch_header:
        page.draw_rect(fitz.Rect(0, 0, width, 22 * mm), color=(1, 1, 1), fill=(1, 1, 1), overlay=True)
        insert_right_text(page, "测试文档", 13 * mm)
        page.draw_line((20 * mm, 16 * mm), (width - 20 * mm, 16 * mm), color=(0.47, 0.47, 0.47), width=0.4, overlay=True)
    page.draw_rect(fitz.Rect(width - 190, height - 105, width - 20, height - 12), color=(1, 1, 1), fill=(1, 1, 1), overlay=True)
    insert_right_text(page, page_text, height - 13 * mm)


def patch_template_footer(page, page_no: int, total_pages: int):
    draw_fitz_header_footer(page, page_no, total_pages, patch_header=False)


def patch_generated_header_footer(page, page_no: int, total_pages: int):
    draw_fitz_header_footer(page, page_no, total_pages, patch_header=True)


def assemble_template_preserved_pdf(template_path: str | os.PathLike, content_path: Path, heading_pages: dict[str, int], materials: list[ProgramMaterial], output_path: str | os.PathLike):
    regular_font, bold_font = register_fonts()
    template_doc = fitz.open(str(template_path))
    content_doc = fitz.open(str(content_path))
    if template_doc.page_count < 3:
        template_doc.close()
        content_doc.close()
        raise ValueError("03-数据统计分析_测试文档模板至少需要包含封面、修订记录和 1-4 章静态内容。")
    static_start, static_end = find_template_static_chapter_range(template_doc)
    static_page_count = static_end - static_start + 1

    if "5. 测试内容" in heading_pages:
        content_start = heading_pages["5. 测试内容"]
    else:
        content_start = find_heading_page(content_doc, "5. 测试内容", start=0, default=6 if content_doc.page_count > 6 else 0)
    toc_page_count = 2
    static_entry_pages = build_static_entry_pages(template_doc, static_start, static_end, toc_page_count)
    prefix_count = 2 + toc_page_count + static_page_count
    generated_page_count = content_doc.page_count - content_start
    total_pages = prefix_count + generated_page_count
    entries, _ = build_toc_entries(materials, content_doc, heading_pages, content_start, toc_page_count, static_page_count, static_entry_pages)

    with tempfile.TemporaryDirectory(prefix="document_filler_toc_") as toc_temp:
        toc_path = Path(toc_temp) / "toc.pdf"
        links, actual_toc_pages = render_toc_pdf(toc_path, entries, total_pages, regular_font, bold_font, min_pages=toc_page_count)
        if actual_toc_pages != toc_page_count:
            toc_page_count = actual_toc_pages
            static_entry_pages = build_static_entry_pages(template_doc, static_start, static_end, toc_page_count)
            prefix_count = 2 + toc_page_count + static_page_count
            total_pages = prefix_count + generated_page_count
            entries, _ = build_toc_entries(materials, content_doc, heading_pages, content_start, toc_page_count, static_page_count, static_entry_pages)
            links, actual_toc_pages = render_toc_pdf(toc_path, entries, total_pages, regular_font, bold_font, min_pages=toc_page_count)

        toc_doc = fitz.open(str(toc_path))
        final_doc = fitz.open()
        final_doc.insert_pdf(template_doc, from_page=0, to_page=1)
        final_doc.insert_pdf(toc_doc)
        final_doc.insert_pdf(template_doc, from_page=static_start, to_page=static_end)
        final_doc.insert_pdf(content_doc, from_page=content_start, to_page=content_doc.page_count - 1)

        preserved_indices = list(range(0, 2)) + list(range(2 + actual_toc_pages, 2 + actual_toc_pages + static_page_count))
        for page_index in preserved_indices:
            patch_template_footer(final_doc[page_index], page_index + 1, final_doc.page_count)
        for page_index in range(2 + actual_toc_pages + static_page_count, final_doc.page_count):
            patch_generated_header_footer(final_doc[page_index], page_index + 1, final_doc.page_count)
        for toc_page_index, rect, target_page in links:
            if target_page < final_doc.page_count:
                final_doc[2 + toc_page_index].insert_link({"kind": fitz.LINK_GOTO, "from": rect, "page": target_page})

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        if output.exists():
            output.unlink()
        final_doc.save(str(output), garbage=4, deflate=True)
        final_doc.close()
        toc_doc.close()
    template_doc.close()
    content_doc.close()


def build_stats_test_pdf(ledger_path: str, service_dir: str, template_path: str, output_path: str):
    """Generate 03-数据统计分析_测试文档 PDF."""
    if template_path and not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="document_filler_stats_test_") as temp_dir:
        work_dir = Path(temp_dir)
        programs, groups, meta = load_ledger_programs(ledger_path, service_dir)
        reports = extract_self_reports(ledger_path, meta, work_dir)
        for order, report in list(reports.items()):
            reports[order] = parse_report_images(report, work_dir)
        materials, summary = build_materials(programs, reports)
        content_path = work_dir / "generated_content.pdf"
        heading_pages = build_pdf(content_path, materials, groups, service_dir)
        assemble_template_preserved_pdf(template_path, content_path, heading_pages, materials, output_path)
    print(json.dumps(summary, ensure_ascii=False))
    print(f"已保存: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Build 03-数据统计分析_测试文档 PDF")
    parser.add_argument("--ledger", required=True)
    parser.add_argument("--service-dir", required=True)
    parser.add_argument("--template", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    build_stats_test_pdf(args.ledger, args.service_dir, args.template, args.output)


if __name__ == "__main__":
    main()
