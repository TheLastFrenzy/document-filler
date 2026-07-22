import importlib.util
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "fill_document.py"
BUILDER_SCRIPT = ROOT / "scripts" / "build_stats_result_usage_workbook.py"


def load_fill_document_module():
    spec = importlib.util.spec_from_file_location("fill_document", SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def load_builder_module():
    spec = importlib.util.spec_from_file_location("build_stats_result_usage_workbook", BUILDER_SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def make_ledger(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["服务目录", "程序XML文本", "统计分析结果表清单"])
    ws.append(["N08-数据统计分析", "<mxGraphModel />", "测试结果表 TEST_RESULT"])
    wb.save(path)


def make_new_column_ledger(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["服务目录", "程序XML文本", "结果表清单"])
    ws.append(["N08-数据统计分析", "<mxGraphModel />", "测试结果表 TEST_RESULT"])
    wb.save(path)


def make_result_template(path: Path):
    wb = openpyxl.Workbook()
    wb.active.title = "说明"
    wb.create_sheet("1、数据源表list")
    wb.create_sheet("2、表融合关系")
    wb.create_sheet("3、数据统计分析结果表list")
    wb.create_sheet("4、数据统计分析结果表详情")
    wb.save(path)


def make_records_with_flowcharts(temp: Path):
    flowcharts = []
    for index in range(1, 3):
        image_path = temp / f"flowchart-{index}.png"
        Image.new("RGB", (16, 16), "white").save(image_path)
        flowcharts.append(image_path)
    return [
        {
            "result_cn": "结果表一",
            "result_en": "RESULT_ONE",
            "sources": ["SRC_ONE"],
            "logic_steps": ["读取源表一", "写入结果表一"],
            "flowchart": str(flowcharts[0]),
            "fields": [
                {
                    "字段中文名": "字段一",
                    "字段英文名": "FIELD_ONE",
                    "字段类型": "VARCHAR2(20)",
                    "默认": "",
                    "不可为空": "No",
                    "唯一": "No",
                    "字段注释": "字段说明",
                }
            ],
        },
        {
            "result_cn": "结果表二",
            "result_en": "RESULT_TWO",
            "sources": ["SRC_TWO"],
            "logic_steps": ["读取源表二", "写入结果表二"],
            "flowchart": str(flowcharts[1]),
            "fields": [
                {
                    "字段中文名": "字段二",
                    "字段英文名": "FIELD_TWO",
                    "字段类型": "NUMBER(10)",
                    "默认": "",
                    "不可为空": "No",
                    "唯一": "No",
                    "字段注释": "字段说明",
                }
            ],
        },
    ]


class StatsResultDispatchTest(unittest.TestCase):
    def test_add_header_row_keeps_generated_excel_headers_center_aligned(self):
        module = load_builder_module()
        wb = openpyxl.Workbook()
        ws = wb.active

        module.add_header_row(ws, ["header 1", "header 2"])
        ws.cell(2, 1, "body")
        module.style_range(ws, 2, 2, 1, 2)

        self.assertEqual(ws["A1"].alignment.horizontal, "center")
        self.assertEqual(ws["B1"].alignment.horizontal, "center")
        self.assertIsNone(ws["A2"].alignment.horizontal)

    def test_business_logic_steps_avoid_unsuitable_xml_logic_terms(self):
        module = load_builder_module()
        record = {
            "result_cn": "热线工单统计结果表",
            "result_en": "FUSION_HOTLINE_ORDER",
            "sources": ["SRC_SECRET"],
            "nodes": [{"sql": "insert into FUSION_HOTLINE_ORDER select ID, AREA_NAME from SRC_SECRET"}],
            "fields": [{"字段中文名": "区域名称"}, {"字段中文名": "工单数量"}],
        }
        resource_info = {
            "SRC_SECRET": {"资源名称": "属地返还加密明细表"},
        }

        steps = module.build_business_logic_steps(record, resource_info)
        text = "\n".join(steps)

        for banned in ["清洗", "抽取", "加密", "质量检查"]:
            self.assertNotIn(banned, text)
        self.assertIn("依据", text)
        self.assertIn("字段口径", text)
        self.assertNotIn("数据范围确认：", text)

    def test_business_logic_steps_use_direct_descriptions_and_concise_write_summary(self):
        module = load_builder_module()
        sql = """
        -- 过滤养老人员表中的无效退休数据
        insert into FUSION_UNION_RETIREMENT
          (member_id, retirement_date, match_state)
        select
          a.member_card_id,
          b.txrq,
          case when b.zjhm is null then '未匹配' else '已匹配' end
        from DWD_TB_UNIONMEMBER a
        left join DWD_ZWY_YLRYQK_A_XXB b
          on a.member_card_id = b.zjhm and a.member_name = b.xm
        where b.jhpt_delete = 0
        """
        record = {
            "result_cn": "工会会员退休信息比对表",
            "result_en": "FUSION_UNION_RETIREMENT",
            "sources": ["DWD_TB_UNIONMEMBER", "DWD_ZWY_YLRYQK_A_XXB"],
            "nodes": [{"label": "退休信息比对", "sql": sql}],
            "fields": [
                {"字段中文名": "会员标识"},
                {"字段中文名": "办理退休日期"},
                {"字段中文名": "匹配状态"},
            ],
        }
        resource_info = {
            "DWD_TB_UNIONMEMBER": {"资源名称": "工会会员信息表"},
            "DWD_ZWY_YLRYQK_A_XXB": {"资源名称": "养老人员情况表"},
        }

        text = "\n".join(module.build_business_logic_steps(record, resource_info))

        self.assertIn("过滤养老人员表中的无效退休数据", text)
        self.assertIn(
            "DWD_TB_UNIONMEMBER.member_card_id = DWD_ZWY_YLRYQK_A_XXB.zjhm",
            text,
        )
        self.assertIn(
            "DWD_TB_UNIONMEMBER.member_name = DWD_ZWY_YLRYQK_A_XXB.xm",
            text,
        )
        self.assertIn("退休信息比对从工会会员信息表（DWD_TB_UNIONMEMBER）读取数据并写入FUSION_UNION_RETIREMENT", text)
        self.assertIn("按条件生成状态、标签或分类结果", text)
        self.assertIn("左外连接", text)
        self.assertIn("关系代数", text)
        self.assertIn("确定性记录链接", text)
        self.assertIn("规则驱动分类", text)
        for prefix in [
            "数据来源准备：",
            "节点业务说明：",
            "节点与来源对应：",
            "表关联：",
            "字段映射（",
            "数据范围确认：",
            "计算方法：",
            "结果字段整理：",
            "结果输出：",
        ]:
            self.assertNotIn(prefix, text)
        self.assertNotIn("<-", text)
        self.assertNotIn("按主键或业务编码关联补齐维度信息", text)
        self.assertNotIn("生成结果表所需的统计口径和明细字段", text)

    def test_business_logic_steps_summarize_generated_and_constant_fields(self):
        module = load_builder_module()
        sql = """
        insert into FUSION_DISTRICT_RESULT (id, area, dsjzx_taskid, org_name)
        select regexp_replace(uuid(), '-', ''), '上海市杨浦区', '${taskid}', org_name
        from MQX_DISTRICT_SOURCE
        where bdc_dt = (select max(bdc_dt) from MQX_DISTRICT_SOURCE)
        """
        record = {
            "result_cn": "区级汇总结果表",
            "result_en": "FUSION_DISTRICT_RESULT",
            "sources": ["MQX_DISTRICT_SOURCE"],
            "nodes": [{"label": "杨浦区", "sql": sql}],
            "fields": [{"字段中文名": "唯一标识"}, {"字段中文名": "所属区"}],
        }

        text = "\n".join(module.build_business_logic_steps(record, {}))

        self.assertIn("杨浦区从MQX_DISTRICT_SOURCE读取数据并写入FUSION_DISTRICT_RESULT", text)
        self.assertIn("生成唯一标识", text)
        self.assertIn("写入任务批次号", text)
        self.assertIn("将所属区固定为“上海市杨浦区”", text)
        self.assertIn("按目标表结构整理业务字段", text)
        self.assertNotIn("<-", text)
        self.assertIn("时态数据最新快照", text)

    def test_business_logic_steps_do_not_invent_database_theory(self):
        module = load_builder_module()
        record = {
            "result_cn": "简单结果表",
            "result_en": "SIMPLE_RESULT",
            "sources": ["SIMPLE_SOURCE"],
            "nodes": [
                {
                    "label": "写入结果",
                    "sql": "insert into SIMPLE_RESULT (id) select id from SIMPLE_SOURCE",
                }
            ],
            "fields": [{"字段中文名": "标识"}],
        }

        text = "\n".join(module.build_business_logic_steps(record, {}))

        for unsupported in ["窗口函数", "集合并运算", "确定性记录链接", "规则驱动分类"]:
            self.assertNotIn(unsupported, text)

    def test_business_logic_steps_keep_every_write_node_without_expanding_mappings(self):
        module = load_builder_module()
        nodes = []
        for index in range(1, 13):
            nodes.append(
                {
                    "label": f"地区{index}",
                    "sql": (
                        "insert into RESULT_TABLE "
                        "(id, area, field_1, field_2, field_3, field_4, field_5, field_6, field_7) "
                        f"select id, '地区{index}', field_1_{index}, field_2_{index}, "
                        f"field_3_{index}, field_4_{index}, field_5_{index}, field_6_{index}, "
                        f"field_7_{index} from SOURCE_{index}"
                    ),
                }
            )
        record = {
            "result_cn": "地区结果表",
            "result_en": "RESULT_TABLE",
            "sources": [f"SOURCE_{index}" for index in range(1, 13)],
            "nodes": nodes,
            "fields": [{"字段中文名": "标识"}, {"字段中文名": "地区"}],
        }

        steps = module.build_business_logic_steps(record, {})
        text = "\n".join(steps)
        write_steps = [line for line in steps if "写入RESULT_TABLE" in line]

        self.assertIn("SOURCE_12", text)
        self.assertIn("地区12从SOURCE_12读取数据并写入RESULT_TABLE", text)
        self.assertEqual(len(write_steps), 12)
        self.assertNotIn("field_7 <- field_7_12", text)
        self.assertNotIn("<-", text)

    def test_relation_row_height_expands_for_detailed_logic(self):
        module = load_builder_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            template = temp / "template.xlsx"
            output = temp / "out.xlsx"
            make_result_template(template)
            records = make_records_with_flowcharts(temp)
            records[0]["logic_steps"] = [
                f"步骤{index}：说明源表字段、目标字段、关联条件和业务规则。" * 3
                for index in range(1, 10)
            ]

            module.build_workbook(template, output, records, {})

            workbook = openpyxl.load_workbook(output)
            relation = workbook["2、表融合关系"]
            height = relation.row_dimensions[2].height
            workbook.close()

        self.assertGreater(height, 220)

    def test_dispatches_to_stats_result_workbook_builder(self):
        module = load_fill_document_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            ledger = temp / "ledger.xlsx"
            template = temp / "template.xlsx"
            catalog = temp / "catalog.xlsx"
            output = temp / "out.xlsx"
            make_ledger(ledger)
            openpyxl.Workbook().save(template)
            openpyxl.Workbook().save(catalog)

            calls = []

            def fake_builder(excel_path, service_dir, template_path, output_path, catalog_path):
                calls.append((excel_path, service_dir, template_path, output_path, catalog_path))
                Path(output_path).write_bytes(b"generated")
                return output_path

            with mock.patch.object(module, "fill_stats_result_usage_workbook", fake_builder, create=True):
                result = module.fill_document(
                    excel_path=str(ledger),
                    service_dir="N08-数据统计分析",
                    material_type="04-数据统计分析_结果表及使用说明",
                    template_path=str(template),
                    output_path=str(output),
                    catalog_path=str(catalog),
                )

            self.assertEqual(result, str(output))
            self.assertEqual(calls, [(str(ledger), "N08-数据统计分析", str(template), str(output), str(catalog))])
            self.assertEqual(output.read_bytes(), b"generated")

    def test_result_workbook_builder_accepts_new_result_list_column_name(self):
        module = load_builder_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            ledger = temp / "ledger.xlsx"
            make_new_column_ledger(ledger)

            rows = module.load_ledger_rows(ledger, "N08-数据统计分析")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["统计分析结果表清单"], "测试结果表 TEST_RESULT")

    def test_draw_flowchart_png_uses_transparent_background_and_no_push_node(self):
        module = load_builder_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            image_path = temp / "flowchart.png"
            record = {
                "row": 1,
                "result_cn": "结果表一",
                "result_en": "RESULT_ONE",
                "sources": ["SRC_ONE", "SRC_TWO"],
                "resource_info": {
                    "SRC_ONE": {"资源名称": "源表一"},
                    "SRC_TWO": {"资源名称": "源表二"},
                },
            }

            module.draw_flowchart_png(record, image_path)

            with Image.open(image_path) as img:
                self.assertEqual(img.mode, "RGBA")
                self.assertEqual(img.getpixel((0, 0))[3], 0)
                region = img.crop((445, 800, 835, 910))
                opaque_pixels = sum(
                    count
                    for count, color in region.getcolors(maxcolors=region.width * region.height)
                    if color[3] > 0
                )

            self.assertLess(opaque_pixels, 4000)

    def test_draw_flowchart_png_does_not_route_merge_line_through_staggered_source_node(self):
        module = load_builder_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            image_path = temp / "flowchart.png"
            record = {
                "row": 1,
                "result_cn": "RESULT_CN",
                "result_en": "RESULT_EN",
                "sources": ["SRC_ONE", "SRC_TWO", "SRC_THREE", "SRC_FOUR", "SRC_FIVE"],
                "resource_info": {},
            }

            module.draw_flowchart_png(record, image_path)

            with Image.open(image_path) as img:
                lower_source_inner = img.crop((510, 395, 770, 475))
                line_pixels = sum(
                    count
                    for count, color in lower_source_inner.getcolors(
                        maxcolors=lower_source_inner.width * lower_source_inner.height
                    )
                    if color == (43, 43, 43, 255)
                )

            self.assertEqual(line_pixels, 0)

    def test_draw_node_keeps_full_long_result_table_name_visible(self):
        module = load_builder_module()
        image = module.Image.new("RGBA", (1280, 1120), (255, 255, 255, 0))
        real_draw = module.ImageDraw.Draw(image)

        class RecordingDraw:
            def __init__(self, delegate):
                self.delegate = delegate
                self.text_lines = []

            def __getattr__(self, name):
                return getattr(self.delegate, name)

            def text(self, position, value, **kwargs):
                self.text_lines.append(value)
                return self.delegate.text(position, value, **kwargs)

        draw = RecordingDraw(real_draw)
        font_file = module.font_path()
        font = module.ImageFont.truetype(font_file, 24) if font_file else module.ImageFont.load_default()
        result_en = "FUSION_QSK_GENERAL_RELATIVE_CODE_FINAL_SHYJSCKK_GA"
        text = (
            "统计融合为\n"
            "亲属关系核验（参考库）个人身后一件事_公安\n"
            f"{result_en}"
        )

        module.draw_node(draw, (435, 510, 845, 680), text, font)

        self.assertIn(result_en, "".join(draw.text_lines))

    def test_result_workbook_applies_songti_gray_fill_defaults_and_relation_spacing(self):
        module = load_builder_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            template = temp / "template.xlsx"
            output = temp / "out.xlsx"
            make_result_template(template)
            records = make_records_with_flowcharts(temp)
            resource_info = {
                "SRC_ONE": {"数据目录代码": "DIR-SRC-1", "资源名称": "源表一", "单位名称": "原始单位一", "资源类型": "原始类型一"},
                "SRC_TWO": {"数据目录代码": "DIR-SRC-2", "资源名称": "源表二", "单位名称": "原始单位二", "资源类型": "原始类型二"},
                "RESULT_ONE": {"数据目录代码": "DIR-RESULT-1", "资源名称": "结果表一", "单位名称": "结果单位一", "资源类型": "结果类型一"},
                "RESULT_TWO": {"资源名称": "结果表二"},
            }

            module.build_workbook(template, output, records, resource_info)

            wb = openpyxl.load_workbook(output)
            source_list = wb["1、数据源表list"]
            relation = wb["2、表融合关系"]
            result_list = wb["3、数据统计分析结果表list"]
            result_detail = wb["4、数据统计分析结果表详情"]

            self.assertEqual(source_list["D2"].value, "上海市大数据中心")
            self.assertEqual(source_list["E2"].value, "库表")
            self.assertEqual(source_list["D3"].value, "上海市大数据中心")
            self.assertEqual(source_list["E3"].value, "库表")
            self.assertEqual(
                [result_list.cell(1, col).value for col in range(1, 7)],
                ["序号", "资源编目（非必填）", "资源名称", "数据提供方", "资源类型", "数据统计分析表的资源信息（表名）"],
            )
            self.assertEqual([result_list.cell(2, col).value for col in range(1, 7)], [1, "DIR-RESULT-1", "结果表一", "结果单位一", "结果类型一", "RESULT_ONE"])
            self.assertEqual([result_list.cell(3, col).value for col in range(1, 7)], [2, None, "结果表二", "上海市大数据中心", "库表", "RESULT_TWO"])
            self.assertNotIn("A1:C1", [str(merged) for merged in relation.merged_cells.ranges])
            self.assertEqual(relation["A1"].value, "2.1 RESULT_ONE")
            self.assertIsNone(relation["B1"].value)
            self.assertIsNone(relation["C1"].value)
            self.assertIsNone(relation["A1"].fill.fill_type)
            self.assertIsNone(relation["B2"].fill.fill_type)
            self.assertIsNone(relation["B3"].fill.fill_type)
            self.assertIsNone(relation["A4"].value)
            self.assertEqual(relation.row_dimensions[4].height, 20)
            self.assertNotIn("A5:C5", [str(merged) for merged in relation.merged_cells.ranges])
            self.assertEqual(relation["A5"].value, "2.2 RESULT_TWO")
            self.assertIsNone(relation["A5"].fill.fill_type)
            self.assertIsNone(relation["B6"].fill.fill_type)
            self.assertIsNone(relation["B7"].fill.fill_type)
            detail_merges = [str(merged) for merged in result_detail.merged_cells.ranges]
            self.assertNotIn("A1:H1", detail_merges)
            self.assertNotIn("B2:H2", detail_merges)
            self.assertEqual(result_detail["A1"].value, "4.1 结果表一")
            self.assertIsNone(result_detail["A1"].fill.fill_type)
            self.assertIsNone(result_detail["B1"].value)
            self.assertEqual(result_detail["B2"].value, "RESULT_ONE")
            self.assertTrue(result_detail["B2"].font.bold)
            self.assertEqual(result_detail["B2"].alignment.horizontal, "center")
            self.assertIsNone(result_detail["C2"].value)
            self.assertNotIn("A6:H6", detail_merges)
            self.assertNotIn("B7:H7", detail_merges)
            self.assertEqual(result_detail["A6"].value, "4.2 结果表二")
            self.assertIsNone(result_detail["A6"].fill.fill_type)
            self.assertEqual(result_detail["B7"].value, "RESULT_TWO")

            for worksheet in wb.worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value not in (None, ""):
                            self.assertEqual(cell.font.name, "宋体")

            for cell in [source_list["A1"], wb["3、数据统计分析结果表list"]["A1"], wb["4、数据统计分析结果表详情"]["B3"]]:
                self.assertEqual(cell.fill.fgColor.rgb, "00F2F2F2")
            wb.close()


if __name__ == "__main__":
    unittest.main()

