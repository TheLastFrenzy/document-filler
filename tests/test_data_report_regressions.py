import importlib.util
import html
import io
import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl
from PIL import Image, ImageDraw
from docx.oxml.ns import qn


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "fill_document.py"
CATALOG_COL = "02-数据报表_设计文档-数据来源库表清单对应数据目录代码"
PNG_1X1 = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
    b"\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00"
    b"\x00\x00\x0cIDATx\x9cc\xf8\xff\xff?\x00\x05\xfe\x02"
    b"\xfeA\xe2!Q\x00\x00\x00\x00IEND\xaeB`\x82"
)


def load_fill_document_module():
    spec = importlib.util.spec_from_file_location("fill_document", SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def content_bbox(path: Path):
    image = Image.open(path).convert("RGB")
    mask = image.point(lambda value: 255 if value < 245 else 0)
    return mask.getbbox(), image.size


class DataReportRegressionTest(unittest.TestCase):
    def test_requirement_text_adds_etc_for_partial_catalog_codes_and_sanitizes_delivery(self):
        module = load_fill_document_module()
        row = {
            "数据需求": "本次需求围绕002420412/000269、002420412/000114目录开展统计核验。",
            "交付要求": "验收时抽查929945317274169及【字段名称】并复核检查结果。",
            CATALOG_COL: "002420412/000269\n002420412/000114\nMB2F30661/000152",
            "工单内容": "数据目录字段统计",
            "业务说明": "根据附件梳理字段名称、数据类型、空值数等内容。",
        }

        catalog_context = {
            "002420412/000269": {
                "资源名称": "重点人员资格表",
                "资源编码": "T_KEY_PERSON",
                "字段": ["姓名", "证件号码", "人员状态"],
            },
            "002420412/000114": {
                "资源名称": "服务机构信息表",
                "资源编码": "T_ORG_INFO",
                "字段": ["机构名称", "所属街镇"],
            },
        }

        normalized = module.normalize_data_report_text_fields(row, catalog_context)

        self.assertIn("002420412/000269、002420412/000114等目录", normalized["数据需求"])
        for banned in ["检查", "抽查", "复核"]:
            self.assertNotIn(banned, normalized["交付要求"])
        self.assertNotIn("929945317274169", normalized["交付要求"])
        self.assertFalse(any(mark in normalized["交付要求"] for mark in "【】[]"))

    def test_requirement_text_uses_catalog_fields_program_fields_and_report_names(self):
        module = load_fill_document_module()
        row = {
            "数据需求": "",
            "交付要求": "",
            CATALOG_COL: "DIR001/000001\nDIR001/000002",
            "工单内容": "便捷共享目录信息统计",
            "业务说明": (
                "根据附件统计目录共享情况。本次工作拟产出以下2份报表成果：\n"
                "便捷共享目录信息表\n目录字段空值统计表"
            ),
            "_programs": [
                {
                    "program_cn": "目录字段空值统计",
                    "program_en": "BGT_DIR_FIELD_NULL",
                    "field_comments": ["目录代码", "目录名称", "空值数"],
                }
            ],
            "_attachment_names": ["便捷共享目录信息表.xlsx"],
        }
        catalog_context = {
            "DIR001/000001": {
                "资源名称": "共享目录基本信息表",
                "资源编码": "T_DIR_INFO",
                "字段": ["目录代码", "目录名称", "提供方名称", "更新频率"],
            },
            "DIR001/000002": {
                "资源名称": "目录字段明细表",
                "资源编码": "T_DIR_FIELD",
                "字段": ["字段中文名", "字段英文名", "空值数"],
            },
        }

        normalized = module.normalize_data_report_text_fields(row, catalog_context)

        self.assertIn("共享目录基本信息表", normalized["数据需求"])
        self.assertIn("目录字段明细表", normalized["数据需求"])
        self.assertIn("目录代码、目录名称、提供方名称", normalized["数据需求"])
        self.assertIn("便捷共享目录信息表", normalized["数据需求"])
        self.assertIn("字段中文名", normalized["交付要求"])
        self.assertIn("目录字段空值统计", normalized["交付要求"])
        self.assertIn("可追溯", normalized["交付要求"])
        self.assertNotEqual(
            normalized["数据需求"],
            "本次需求围绕DIR001/000001、DIR001/000002目录开展统计整理，结合便捷共享目录信息统计确认字段口径、统计范围和结果呈现内容。",
        )

    def test_requirement_document_dispatch_builds_catalog_context_when_catalog_is_supplied(self):
        module = load_fill_document_module()
        rows = [{"工单内容": "目录统计"}]
        context = {"DIR001/000001": {"资源名称": "共享目录基本信息表"}}

        with mock.patch.object(module, "read_excel", return_value=rows) as read_excel:
            with mock.patch.object(module, "build_data_report_catalog_context", return_value=context) as build_context:
                with mock.patch.object(module, "fill_requirement_doc", return_value="out.docx") as fill_requirement:
                    result = module.fill_document(
                        excel_path="ledger.xlsx",
                        service_dir="N08-数据报表服务",
                        material_type="01-数据报表_需求文档",
                        template_path="template.docx",
                        output_path="out.docx",
                        catalog_path="catalog.xlsx",
                    )

        self.assertEqual(result, "out.docx")
        read_excel.assert_called_once_with("ledger.xlsx", "N08-数据报表服务")
        build_context.assert_called_once_with("catalog.xlsx", rows)
        fill_requirement.assert_called_once_with(rows, "template.docx", "out.docx", catalog_context=context)

    def test_requirement_report_names_detect_marker_without_yixia(self):
        module = load_fill_document_module()

        text = module.report_names_from_business_text(
            {
                "业务说明": "根据附件统计目录共享情况。本次工作拟产出2份报表成果：\n目录共享表\n空值统计表",
            }
        )

        self.assertIn("目录共享表", text)
        self.assertIn("空值统计表", text)

    def test_requirement_report_names_fall_back_to_result_program_list(self):
        module = load_fill_document_module()

        text = module.report_names_from_business_text(
            {
                "业务说明": "根据附件统计目录共享情况。",
                "统计分析结果表清单": "目录字段空值统计 BGT_DIR_FIELD_NULL",
            }
        )

        self.assertIn("目录字段空值统计", text)
        self.assertNotEqual(text, "报表成果")

    def test_requirement_and_delivery_endings_vary_and_delivery_names_output_formats(self):
        module = load_fill_document_module()
        catalog_context = {
            "DIR001/000001": {
                "资源名称": "共享目录基本信息表",
                "资源编码": "T_DIR_INFO",
                "字段": ["目录代码", "目录名称", "提供方名称", "更新频率"],
            },
            "DIR001/000002": {
                "资源名称": "目录字段明细表",
                "资源编码": "T_DIR_FIELD",
                "字段": ["字段中文名", "字段英文名", "空值数"],
            },
        }
        rows = [
            {
                "工单内容": "便捷共享目录信息统计",
                "业务说明": "根据附件统计目录共享情况。",
                CATALOG_COL: "DIR001/000001\nDIR001/000002",
                "统计分析结果表清单": "目录共享表 BGT_DIR_SHARE",
                "_attachment_names": ["目录共享表.xlsx"],
            },
            {
                "工单内容": "下发量统计指标逻辑变更",
                "业务说明": "根据附件调整下发量统计口径。",
                CATALOG_COL: "DIR001/000001\nDIR001/000002",
                "统计分析结果表清单": "下发量统计表 BGT_SEND_COUNT",
                "_attachment_names": ["下发量统计表.xlsx", "口径说明.docx"],
            },
            {
                "工单内容": "接口调用风险明细统计",
                "业务说明": "根据附件梳理接口调用异常明细。",
                CATALOG_COL: "DIR001/000001\nDIR001/000002",
                "统计分析结果表清单": "接口风险明细表 BGT_API_RISK",
            },
        ]

        normalized = [module.normalize_data_report_text_fields(row, catalog_context) for row in rows]
        requirement_endings = {module.final_cn_sentence(item["数据需求"]) for item in normalized}
        delivery_endings = {module.final_cn_sentence(item["交付要求"]) for item in normalized}

        self.assertGreater(len(requirement_endings), 1)
        self.assertGreater(len(delivery_endings), 1)
        self.assertIn("Excel电子表格报表", normalized[0]["交付要求"])
        self.assertNotIn("Word文档说明材料", normalized[0]["交付要求"])
        self.assertIn("Excel电子表格报表", normalized[1]["交付要求"])
        self.assertIn("Word文档说明材料", normalized[1]["交付要求"])
        self.assertNotIn("Excel电子表格报表", normalized[2]["交付要求"])
        self.assertNotIn("Word文档说明材料", normalized[2]["交付要求"])
        for item in normalized:
            delivery = item["交付要求"]
            self.assertNotIn("交付物后缀显示", delivery)
            self.assertNotIn("格式提交", delivery)
            self.assertIn("命名", delivery)
            self.assertIn("统计时间段", delivery)

    def test_delivery_format_wording_uses_actual_attachment_suffixes_only(self):
        module = load_fill_document_module()
        base_row = {
            "工单内容": "接口调用风险明细统计",
            "业务说明": "根据附件梳理接口调用异常明细。",
            CATALOG_COL: "DIR001/000001",
        }

        excel_row = dict(base_row, _attachment_names=["接口风险明细表.xlsx"])
        mixed_row = dict(base_row, _attachment_names=["接口风险明细表.xlsx", "分析说明.docx"])
        pdf_row = dict(base_row, 交付物="接口风险处置清单.pdf")
        no_suffix_row = dict(base_row)

        excel_delivery = module.normalize_data_report_text_fields(excel_row)["交付要求"]
        mixed_delivery = module.normalize_data_report_text_fields(mixed_row)["交付要求"]
        pdf_delivery = module.normalize_data_report_text_fields(pdf_row)["交付要求"]
        no_suffix_delivery = module.normalize_data_report_text_fields(no_suffix_row)["交付要求"]

        self.assertIn("Excel电子表格报表", excel_delivery)
        self.assertNotIn("Word文档说明材料", excel_delivery)
        self.assertNotIn("PDF版定稿或签收材料", excel_delivery)
        self.assertIn("Excel电子表格报表", mixed_delivery)
        self.assertIn("Word文档说明材料", mixed_delivery)
        self.assertIn("PDF版定稿或签收材料", pdf_delivery)
        self.assertNotIn("Excel电子表格报表", no_suffix_delivery)
        self.assertNotIn("Word文档说明材料", no_suffix_delivery)
        self.assertIn("交付材料按实际附件内容整理", no_suffix_delivery)
        for delivery in [excel_delivery, mixed_delivery, pdf_delivery, no_suffix_delivery]:
            self.assertNotIn("交付物后缀显示", delivery)
            self.assertNotIn("格式提交", delivery)

    def test_delivery_format_ignores_generated_internal_zip_wrappers(self):
        module = load_fill_document_module()
        row = {
            "工单内容": "涉企数据资源目录梳理",
            "业务说明": "根据附件梳理涉企数据目录字段情况。",
            CATALOG_COL: "DIR001/000001",
            "_attachment_files": [Path("Workbook.xls"), Path("Workbook_embedded.zip")],
        }

        delivery = module.normalize_data_report_text_fields(row)["交付要求"]

        self.assertIn("Excel电子表格报表", delivery)
        self.assertNotIn("压缩包", delivery)

    def test_requirement_document_dispatch_attaches_delivery_file_names_for_format_inference(self):
        module = load_fill_document_module()
        rows = [{"_row": 2, "工单内容": "目录统计"}]

        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = Path(temp_dir) / "ledger.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["服务目录", "交付物"])
            ws.append(["N08-数据报表服务", ""])
            wb.save(ledger)

            attachments = {2: [Path("目录共享表.xlsx"), Path("口径说明.docx")]}
            with mock.patch.object(module, "read_excel", return_value=rows):
                with mock.patch.object(module, "extract_deliverable_attachments", return_value=attachments) as extract:
                    with mock.patch.object(module, "fill_requirement_doc", return_value="out.docx") as fill_requirement:
                        result = module.fill_document(
                            excel_path=str(ledger),
                            service_dir="N08-数据报表服务",
                            material_type="01-数据报表_需求文档",
                            template_path="template.docx",
                            output_path="out.docx",
                        )

        self.assertEqual(result, "out.docx")
        extract.assert_called_once()
        passed_rows = fill_requirement.call_args.args[0]
        self.assertEqual(passed_rows[0]["_attachment_files"], attachments[2])
        self.assertEqual(passed_rows[0]["_attachment_names"], ["目录共享表.xlsx", "口径说明.docx"])

    def test_select_excel_preview_range_prefers_richer_sheet_and_caps_wide_ranges(self):
        module = load_fill_document_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "book.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "empty"
            ws["A1"] = "only one cell"
            rich = wb.create_sheet("rich")
            for row in range(1, 21):
                for col in range(1, 16):
                    rich.cell(row, col).value = f"R{row}C{col}"
            wb.save(path)

            sheet_name, first_row, first_col, last_row, last_col = module.select_excel_preview_range(path)

        self.assertEqual(sheet_name, "rich")
        self.assertEqual((first_row, first_col), (1, 1))
        self.assertEqual(last_row, 20)
        self.assertEqual(last_col - first_col + 1, 9)

    def test_crop_and_normalize_image_removes_large_blank_margins_and_keeps_resolution(self):
        module = load_fill_document_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "preview.png"
            image = Image.new("RGB", (1000, 800), "white")
            draw = ImageDraw.Draw(image)
            draw.rectangle((350, 280, 650, 420), fill="black")
            image.save(path)

            self.assertTrue(module.crop_and_normalize_image(path, min_width=1200, min_height=500))
            bbox, size = content_bbox(path)

        self.assertIsNotNone(bbox)
        width, height = size
        left, top, right, bottom = bbox
        self.assertGreaterEqual(width, 1200)
        self.assertGreaterEqual(height, 500)
        max_margin = max(left / width, top / height, (width - right) / width, (height - bottom) / height)
        self.assertLess(max_margin, 0.18)

    def test_image_bytes_for_docx_converts_pillow_readable_images_to_png(self):
        module = load_fill_document_module()
        source = io.BytesIO()
        Image.new("RGB", (12, 8), "red").save(source, format="JPEG")

        converted = module.image_bytes_for_docx(source.getvalue())

        self.assertTrue(converted.startswith(b"\x89PNG\r\n\x1a\n"))

    def test_attachment_preview_uses_excel_range_rendering_before_pdf_fallback(self):
        module = load_fill_document_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            source = temp / "attachment.xlsx"
            openpyxl.Workbook().save(source)
            calls = []

            def fake_excel_range_to_png(excel_path, image_path):
                calls.append((excel_path, image_path))
                image_path.write_bytes(PNG_1X1)
                return True

            with mock.patch.object(module, "excel_range_to_png", fake_excel_range_to_png, create=True):
                preview = module.generate_attachment_screenshot_bytes([source], temp, row_number=2)

        self.assertEqual(preview, PNG_1X1)
        self.assertEqual(len(calls), 1)
        self.assertEqual(calls[0][0], source)

    def test_legacy_xls_attachment_skips_openpyxl_range_rendering(self):
        module = load_fill_document_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            source = temp / "legacy.xls"
            source.write_bytes(b"legacy xls placeholder")

            def fail_excel_range_to_png(*_args, **_kwargs):
                raise AssertionError(".xls must not use openpyxl range rendering")

            def fake_office_export_to_pdf(_source, pdf_path):
                pdf_path.write_bytes(b"%PDF-1.4\n%fake")
                return True

            def fake_render_pdf_first_page(_pdf_path, image_path):
                image_path.write_bytes(PNG_1X1)
                return True

            with mock.patch.object(module, "excel_range_to_png", fail_excel_range_to_png):
                with mock.patch.object(module, "office_export_to_pdf", fake_office_export_to_pdf):
                    with mock.patch.object(module, "render_pdf_first_page", fake_render_pdf_first_page):
                        preview = module.generate_attachment_screenshot_bytes([source], temp, row_number=2)

        self.assertEqual(preview, PNG_1X1)

    def test_read_data_report_design_groups_merges_work_order_program_rows(self):
        module = load_fill_document_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            ledger = Path(temp_dir) / "ledger.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = [
                "服务目录",
                "需求单号",
                "工单号",
                "工单内容",
                CATALOG_COL,
                "统计分析结果表清单",
                "程序XML文本",
                "交付物",
            ]
            ws.append(headers)
            ws.append([
                "N08-数据报表服务",
                "REQ-1",
                "WO-1",
                "合并工单",
                "DIR-1",
                "程序甲 A1",
                "<xml>a1</xml>",
                "",
            ])
            ws.append([
                None,
                None,
                None,
                None,
                None,
                "程序乙 A2",
                "<xml>a2</xml>",
                None,
            ])
            for col in range(1, 6):
                ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            ws.merge_cells(start_row=2, start_column=8, end_row=3, end_column=8)
            wb.save(ledger)

            groups = module.read_data_report_design_groups(str(ledger), "N08-数据报表服务")

        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0]["_row_numbers"], [2, 3])
        self.assertEqual(
            [(p["program_cn"], p["program_en"], p["xml"]) for p in groups[0]["_programs"]],
            [("程序甲", "A1", "<xml>a1</xml>"), ("程序乙", "A2", "<xml>a2</xml>")],
        )

    def test_extracts_indicator_field_comments_from_program_xml(self):
        module = load_fill_document_module()
        sql = """
        CREATE TABLE IF NOT EXISTS bgt_fusion.A1(
          id varchar2(20) DEFAULT NULL COMMENT '证件号码',
          name varchar2(50) DEFAULT NULL COMMENT '姓名'
        )
        """
        model_data = html.escape(json.dumps({"stepLabel": "建表", "sql": sql}, ensure_ascii=False), quote=True)
        xml_text = f'<mxGraphModel><root><mxCell id="2" modelData="{model_data}" /></root></mxGraphModel>'

        comments = module.extract_indicator_field_comments_from_program(
            {"program_cn": "程序甲", "program_en": "A1", "xml": xml_text}
        )

        self.assertEqual(comments, ["证件号码", "姓名"])

    def test_extracts_indicator_field_comments_from_truncated_escaped_xml(self):
        module = load_fill_document_module()
        xml_text = (
            '<mxCell id="22" modelData="{&quot;sql&quot;:&quot;'
            'CREATE TABLE IF NOT EXISTS bgt_fusion.${cnt.hb}(\\r\\n'
            'xh varchar2(10) DEFAULT NULL COMMENT &#39;报表序号&#39;,\\r\\n'
            'sjml string DEFAULT NULL COMMENT &#39;数据目录&#39;\\r\\n'
            ')&quot;}"'
        )

        comments = module.extract_indicator_field_comments_from_program(
            {"program_cn": "程序甲", "program_en": "FUSION_SJJ_ZBYC_12YGZHB_1_ONCE", "xml": xml_text}
        )

        self.assertEqual(comments, ["报表序号", "数据目录"])

    def test_extracts_indicator_fields_from_insert_column_list_when_create_table_is_missing(self):
        module = load_fill_document_module()
        sql = (
            "insert into ${resultSourceTable}"
            "(cata_code,cata_title,provider_name,cloumn_comments,data_count)"
            " values('${mldm}','${mlmc}','${tgfmc}','${colums[j].comment}','${var.counts}')"
        )
        model_data = html.escape(json.dumps({"sql": sql}, ensure_ascii=False), quote=True)
        xml_text = f'<mxGraphModel><root><mxCell id="2" modelData="{model_data}" /></root></mxGraphModel>'

        comments = module.extract_indicator_field_comments_from_program(
            {"program_cn": "程序甲", "program_en": "A1", "xml": xml_text}
        )

        self.assertEqual(comments, ["目录代码", "目录名称", "提供方名称", "字段注释", "数据量"])

    def test_indicator_rows_bind_program_fields_to_attachment_names(self):
        module = load_fill_document_module()
        rows = module.build_data_report_indicator_rows(
            [
                {"program_cn": "程序甲", "program_en": "A1", "field_comments": ["证件号码", "姓名"]},
                {"program_cn": "程序乙", "program_en": "A2", "field_comments": ["性别", "年龄"]},
            ],
            ["b1.docx", "b2.xlsx"],
        )

        self.assertEqual(
            rows,
            [
                ["b1", "证件号码", "证件号码", "无", "文本", "无", "证件号码"],
                ["b1", "姓名", "姓名", "无", "文本", "无", "姓名"],
                ["b2", "性别", "性别", "无", "文本", "无", "性别"],
                ["b2", "年龄", "年龄", "无", "文本", "无", "年龄"],
            ],
        )

    def test_generated_attachment_fallback_names_are_not_used_as_report_names(self):
        module = load_fill_document_module()

        for value in [
            "deliverable_row03.xlsx",
            "package.docx",
            "Workbook.xls",
            "Workbook_embedded.zip",
            "Ole10Native_embedded.zip",
        ]:
            self.assertEqual(module.clean_attachment_report_name(value), "")

        rows = module.build_data_report_indicator_rows(
            [{"program_cn": "便捷共享目录信息统计", "program_en": "BGT_REPORT", "field_comments": ["目录名称"]}],
            ["deliverable_row03.xlsx"],
        )

        self.assertEqual(rows[0][0], "便捷共享目录信息统计")
        self.assertNotIn("deliverable_row", rows[0][0])

    def test_result_form_uses_program_names_when_attachment_names_are_generated_fallbacks(self):
        module = load_fill_document_module()

        text = module.infer_data_report_result_form(
            {
                "工单内容": "便捷共享目录信息统计",
                "统计分析结果表清单": "便捷共享目录信息统计 BGT_REPORT",
                "_attachment_names": ["deliverable_row03.xlsx"],
            }
        )

        self.assertIn("便捷共享目录信息统计", text)
        self.assertNotIn("deliverable_row", text)

    def test_office_export_success_accepts_created_output_even_with_nonzero_returncode(self):
        module = load_fill_document_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "converted.pdf"
            target.write_bytes(b"%PDF-1.4\n")

            self.assertTrue(module.office_export_created_output(1, target))

    def test_indicator_rows_round_robin_programs_when_programs_exceed_files(self):
        module = load_fill_document_module()
        rows = module.build_data_report_indicator_rows(
            [
                {"program_cn": "程序甲", "program_en": "A1", "field_comments": ["字段1"]},
                {"program_cn": "程序乙", "program_en": "A2", "field_comments": ["字段2"]},
                {"program_cn": "程序丙", "program_en": "A3", "field_comments": ["字段3"]},
            ],
            ["b1.xlsx", "b2.xlsx"],
        )

        self.assertEqual([row[0] for row in rows], ["b1", "b2", "b1"])
        self.assertEqual({row[0] for row in rows}, {"b1", "b2"})

    def test_design_text_normalization_fills_new_ledger_fallback_fields(self):
        module = load_fill_document_module()
        row = {
            "工单内容": "重点人员专题报表",
            "业务说明": "根据附件统计重点人员数量、行政区划和更新时间。",
            "业务描述": "用于支撑重点人员专题分析。",
            CATALOG_COL: "002420412/000269\n002420412/000114\nMB2F30661/000152",
            "统计分析结果表清单": "重点人员明细 BGT_REPORT_PERSON\n重点人员汇总 BGT_REPORT_SUM",
            "_programs": [
                {
                    "program_cn": "重点人员明细",
                    "program_en": "BGT_REPORT_PERSON",
                    "field_comments": ["姓名", "证件号码"],
                }
            ],
            "_attachment_names": ["重点人员明细.xlsx"],
        }

        catalog_context = {
            "002420412/000269": {
                "资源名称": "重点人员资格表",
                "资源编码": "T_KEY_PERSON",
                "字段": ["姓名", "证件号码", "人员状态"],
            },
            "002420412/000114": {
                "资源名称": "服务机构信息表",
                "资源编码": "T_ORG_INFO",
                "字段": ["机构名称", "所属街镇"],
            },
        }

        normalized = module.normalize_data_report_text_fields(row, catalog_context)

        for key in ("内容描述", "业务场景", "结果形式", "数据处理逻辑"):
            self.assertTrue(module.compact_spaces(normalized.get(key)), key)
        self.assertIn("重点人员", normalized["内容描述"])
        self.assertIn("重点人员", normalized["业务场景"])
        self.assertIn("重点人员资格表", normalized["数据内容"])
        self.assertIn("姓名", normalized["数据内容"])
        self.assertIn("重点人员明细", normalized["结果形式"])
        self.assertIn("002420412/000269", normalized["数据处理逻辑"])
        self.assertIn("姓名", normalized["数据处理逻辑"])
        for text in (normalized["业务场景"], normalized["数据内容"]):
            for banned in ("本节", "附件中可见", "上线后用于", "展示附件", "统计结果查看、业务核验和材料归档"):
                self.assertNotIn(banned, text)

    def test_design_business_scene_and_data_content_are_not_fixed_template_sentences(self):
        module = load_fill_document_module()
        catalog_context = {
            "DIR001/000001": {
                "资源名称": "共享目录基本信息表",
                "资源编码": "T_DIR_INFO",
                "字段": ["目录代码", "目录名称", "下发量"],
            },
            "DIR001/000002": {
                "资源名称": "目录字段明细表",
                "资源编码": "T_DIR_FIELD",
                "字段": ["字段中文名", "字段英文名", "空值数"],
            },
        }
        rows = [
            {
                "工单内容": "便捷共享目录信息统计",
                "业务说明": "根据专题会议要求，编制便捷共享目录信息统计报表。",
                CATALOG_COL: "DIR001/000001\nDIR001/000002",
                "统计分析结果表清单": "目录共享表 BGT_DIR_SHARE",
                "_attachment_names": ["目录共享表.xlsx"],
            },
            {
                "工单内容": "涉企数据资源高价值目录梳理",
                "业务说明": "对涉企数据资源进行高价值目录梳理，形成字段申请次数和属地字段情况统计。",
                CATALOG_COL: "DIR001/000001\nDIR001/000002",
                "统计分析结果表清单": "涉企目录字段统计 BGT_COMPANY_FIELD",
                "_attachment_names": ["涉企目录字段统计.xls"],
            },
            {
                "工单内容": "养老迁入迁出比对",
                "业务说明": "比对养老迁入迁出人员情况，为民政养老业务提供核对清单。",
                CATALOG_COL: "DIR001/000001\nDIR001/000002",
                "统计分析结果表清单": "养老迁入迁出比对表 BGT_OLD_MIGRATE",
                "_attachment_names": ["养老迁入迁出比对表.xlsx"],
            },
        ]

        normalized = [module.normalize_data_report_text_fields(row, catalog_context) for row in rows]
        business_scenes = [item["业务场景"] for item in normalized]
        data_contents = [item["数据内容"] for item in normalized]

        self.assertGreater(len({module.final_cn_sentence(text) for text in business_scenes}), 1)
        self.assertGreater(len({module.final_cn_sentence(text) for text in data_contents}), 1)
        for text in business_scenes + data_contents:
            for banned in ("本节", "附件中可见", "上线后用于", "展示附件", "统计结果查看、业务核验和材料归档"):
                self.assertNotIn(banned, text)
        self.assertIn("共享目录基本信息表", data_contents[0])
        self.assertIn("涉企数据资源", business_scenes[1])
        self.assertIn("养老迁入迁出", business_scenes[2])

    def test_design_business_scene_skips_long_list_like_source_fragments(self):
        module = load_fill_document_module()
        row = {
            "工单内容": "民生社会事业报表梳理",
            "业务说明": (
                "根据市数据局要求，对民生、社会事业领域报表情况进行梳理，包括但不限于1、社会事业领域："
                "区、下发部门、报表名称、报送层级、下发部门确认是否保留、报表链接、下发部门确认是否保留、\"系。"
            ),
            CATALOG_COL: "DIR001/000001",
            "统计分析结果表清单": "民生报表梳理 BGT_LIVELIHOOD_REPORT",
        }

        normalized = module.normalize_data_report_text_fields(row)

        self.assertIn("民生社会事业报表梳理", normalized["业务场景"])
        self.assertIn("民生社会事业报表梳理", normalized["内容描述"])
        for banned in ("包括但不限于1、", "报表链接", "\"系"):
            self.assertNotIn(banned, normalized["业务场景"])
            self.assertNotIn(banned, normalized["内容描述"])

    def test_launch_requirement_description_falls_back_to_new_ledger_business_fields(self):
        module = load_fill_document_module()

        self.assertEqual(
            module.build_launch_requirement_description(
                {
                    "需求描述": "",
                    "业务说明": "根据附件生成重点人员统计报表。",
                    "业务描述": "用于专题分析。",
                    "工单内容": "重点人员专题报表",
                }
            ),
            "根据附件生成重点人员统计报表。",
        )
        self.assertEqual(
            module.build_launch_requirement_description(
                {
                    "需求描述": "保留已有需求描述。",
                    "业务说明": "不应覆盖。",
                }
            ),
            "保留已有需求描述。",
        )
        self.assertIn(
            "空值率统计",
            module.build_launch_requirement_description({"工单内容": "空值率统计"}),
        )

    def test_launch_image_columns_only_returns_available_cells(self):
        module = load_fill_document_module()

        columns = module.launch_image_columns_for_row(
            {"_row": 5},
            {
                (5, "上线交付截图2"): b"image",
                (6, "上线交付截图1"): b"other-row",
            },
            ["上线交付截图1", "上线交付截图2"],
        )

        self.assertEqual(columns, ["上线交付截图2"])

    def test_launch_record_helpers_apply_fixed_id_format_indent_and_word_wrap(self):
        module = load_fill_document_module()
        row = {"需求单号": "X_A_RCGZ_202601070119", "工单号": "G_A_RCGZ_202601070180"}

        self.assertEqual(
            module.build_launch_identifier_line(row),
            "需求编号：X_A_RCGZ_202601070119\t对应工单编号：G_A_RCGZ_202601070180",
        )

        paragraph = module.mp("统计报表：3次。", "Normal", 480, word_wrap=True)
        p_pr = paragraph.find(qn("w:pPr"))
        self.assertIsNotNone(p_pr.find(qn("w:wordWrap")))
        indent = p_pr.find(qn("w:ind"))
        self.assertEqual(indent.get(qn("w:firstLine")), "480")


if __name__ == "__main__":
    unittest.main()
